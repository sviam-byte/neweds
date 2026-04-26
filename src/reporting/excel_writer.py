"""
Генератор Excel-отчетов (локализованный).
"""

from __future__ import annotations

import logging
import os
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from src.config import is_directed_method, is_pvalue_method
from src.core.results import AnalysisResult, tool_adapter_from_result
from src.visualization import plots


def write_excel_report(
    result: AnalysisResult,
    out_dir: str,
    *,
    filename: str = "report.xlsx",
    **kwargs,
) -> str:
    """Write Excel report from public AnalysisResult contract."""

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    tool = tool_adapter_from_result(result)
    return ExcelReportWriter(tool).write(str(out / filename), **kwargs)



@dataclass(slots=True)
class ExcelReportWriter:
    """Класс для записи результатов анализа в Excel файл."""

    tool: object

    EXCEL_MAX_ROWS = 1_048_576
    EXCEL_MAX_COLS = 16_384

    def _safe_sheet_title(self, title: str) -> str:
        """Нормализует title по правилам Excel (запрещенные символы + длина)."""
        title = "".join("_" if ch in '[]:*?/\\' else ch for ch in str(title or "Sheet"))
        title = title.strip() or "Sheet"
        return title[:31]

    def _sheet_can_hold(self, df: pd.DataFrame) -> bool:
        """Проверяет, помещается ли DataFrame в лимиты листа Excel."""
        rows, cols = getattr(df, "shape", (0, 0))
        return int(rows) + 1 <= self.EXCEL_MAX_ROWS and int(cols) <= self.EXCEL_MAX_COLS

    def _add_image_to_sheet(self, ws, buf, cell: str, w: int = 400, h: int = 300) -> None:
        img = Image(buf)
        img.width = w
        img.height = h
        ws.add_image(img, cell)

    def write(self, save_path: str, **kwargs) -> str:
        """Создает и сохраняет Excel-файл со всеми результатами."""
        wb = Workbook()
        warnings: list[str] = []

        ws_data = wb.active
        ws_data.title = "Исходные данные"
        data_df = getattr(self.tool, "data", pd.DataFrame())
        if not isinstance(data_df, pd.DataFrame):
            data_df = pd.DataFrame(data_df)

        if self._sheet_can_hold(data_df):
            ws_data.append(list(data_df.columns))
            for row in dataframe_to_rows(data_df, index=False, header=False):
                ws_data.append(row)
        else:
            preview_rows = min(int(len(data_df)), 10_000)
            preview = data_df.head(preview_rows)
            ws_data.append(["warning", "dataset_too_large_for_single_excel_sheet"])
            ws_data.append(["shape", f"{data_df.shape[0]}x{data_df.shape[1]}"])
            ws_data.append(["preview_rows", preview_rows])
            ws_data.append([])
            ws_data.append(list(preview.columns))
            for row in dataframe_to_rows(preview, index=False, header=False):
                ws_data.append(row)
            warnings.append(
                f"Исходные данные не помещаются в один Excel sheet: shape={data_df.shape}. Сохранён только preview из {preview_rows} строк."
            )

        threshold = kwargs.get("threshold", 0.2)
        p_alpha = kwargs.get("p_value_alpha", 0.05)
        include_ar_diagnostics = bool(kwargs.get("include_ar_diagnostics", True))
        disable_images = str(os.getenv("TS_TOOL_DISABLE_EXCEL_IMAGES", "0")).strip().lower() in {"1", "true", "yes", "on"}
        max_image_n = int(kwargs.get("excel_max_image_n", os.getenv("TS_TOOL_EXCEL_MAX_IMAGE_N", 300) or 300))
        max_matrix_sheet_n = int(kwargs.get("excel_max_matrix_sheet_n", os.getenv("TS_TOOL_EXCEL_MAX_MATRIX_N", 5000) or 5000))

        for variant, mat in (getattr(self.tool, "results", {}) or {}).items():
            if mat is None:
                continue

            ws = wb.create_sheet(title=self._safe_sheet_title(variant))
            ws.append([f"Метод: {variant}"])

            is_pval = is_pvalue_method(variant)
            thr = p_alpha if is_pval else threshold

            arr = np.asarray(mat)
            if arr.ndim != 2 or arr.shape[0] != arr.shape[1]:
                ws.append(["warning", "matrix_is_not_square_or_invalid"])
                warnings.append(f"{variant}: матрица не квадратная или некорректная, лист записан без графиков.")
                continue

            n = int(arr.shape[0])
            labels = list(getattr(data_df, "columns", []))
            if len(labels) != n:
                labels = [f"v{i:04d}" for i in range(n)]

            if n <= max_matrix_sheet_n:
                ws.append(["Матрица значений:"])
                df_mat = pd.DataFrame(arr, columns=labels, index=labels)
                for row in dataframe_to_rows(df_mat, index=True, header=True):
                    ws.append(row)
                last_row = ws.max_row + 2
            else:
                ws.append(["warning", f"matrix_too_large_for_excel_sheet: n={n}, limit={max_matrix_sheet_n}"])
                ws.append(["shape", f"{n}x{n}"])
                last_row = ws.max_row + 2
                warnings.append(f"{variant}: матрица {n}x{n} не выгружена на лист целиком.")

            if disable_images:
                ws.append(["images", "disabled_by_env"])
                continue
            if n > max_image_n:
                ws.append(["images", f"skipped_large_matrix_n={n}"])
                warnings.append(f"{variant}: графики в Excel пропущены для большой матрицы n={n}.")
                continue

            try:
                buf_heat = plots.plot_heatmap(arr, f"{variant} Теплокарта")
                self._add_image_to_sheet(ws, buf_heat, f"A{last_row}")
            except Exception as exc:
                ws.append(["heatmap_error", str(exc)])
                warnings.append(f"{variant}: не удалось построить heatmap: {exc}")

            try:
                buf_conn = plots.plot_connectome(
                    arr,
                    f"{variant} Граф",
                    threshold=thr,
                    directed=is_directed_method(variant),
                    invert_threshold=is_pval,
                )
                self._add_image_to_sheet(ws, buf_conn, f"G{last_row}")
            except Exception as exc:
                ws.append(["connectome_error", str(exc)])
                warnings.append(f"{variant}: не удалось построить connectome: {exc}")


        # AR-диагностика в отдельные листы (если доступна в preprocessing_report.notes).
        if include_ar_diagnostics:
            try:
                ar_df = pd.DataFrame()
                if hasattr(self.tool, "_build_ar_diagnostics_dataframe"):
                    ar_df = self.tool._build_ar_diagnostics_dataframe()
                if isinstance(ar_df, pd.DataFrame) and not ar_df.empty:
                    ws_ar = wb.create_sheet(title="AR_Diagnostics")
                    ws_ar.append(list(ar_df.columns))
                    for row in dataframe_to_rows(ar_df, index=False, header=False):
                        ws_ar.append(row)
                    ws_ar.freeze_panes = "A2"
                    for idx, col in enumerate(ar_df.columns, start=1):
                        sample = [len(str(col))]
                        sample.extend(len(str(v)) for v in ar_df[col].head(200).tolist())
                        ws_ar.column_dimensions[get_column_letter(idx)].width = min(max(12, max(sample) + 2), 28)

                rep = getattr(self.tool, "preprocessing_report", None)
                notes = getattr(rep, "notes", {}) if rep is not None else {}
                ac = (notes or {}).get("autocorr") or {}
                ex = ac.get("examples") or {}
                ex_rows = []
                if isinstance(ex, dict):
                    for name, item in ex.items():
                        if not isinstance(item, dict):
                            continue
                        row = {"series": name, "phi1": item.get("phi1")}
                        for k, v in (item.get("before_corr_by_lag") or {}).items():
                            row[f"before_{k}"] = v
                        for k, v in (item.get("after_corr_by_lag") or {}).items():
                            row[f"after_{k}"] = v
                        ex_rows.append(row)
                if ex_rows:
                    df_ex = pd.DataFrame(ex_rows)
                    ws_ex = wb.create_sheet(title="AR_Examples")
                    ws_ex.append(list(df_ex.columns))
                    for row in dataframe_to_rows(df_ex, index=False, header=False):
                        ws_ex.append(row)
                    ws_ex.freeze_panes = "A2"
            except Exception as exc:
                warnings.append(f"AR diagnostics sheets skipped: {exc}")

        if warnings:
            ws_warn = wb.create_sheet(title="Warnings")
            ws_warn.append(["message"])
            for msg in warnings:
                ws_warn.append([msg])

        wb.save(save_path)
        logging.info("[Excel] Отчет сохранен: %s", save_path)
        return save_path
