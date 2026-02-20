#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Главный движок анализа временных рядов.
Содержит класс BigMasterTool и все функции расчета метрик.

TODO: Этот файл требует дальнейшего рефакторинга - разнести метрики по отдельным модулям.
"""

import argparse
import base64
import datetime as _dt
import html as _html
import importlib
import importlib.util
import logging
import os
import shutil
import warnings
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from itertools import chain, combinations, permutations
from typing import Dict, List, Optional, Tuple

from src.analysis.dimred import apply_dimred

import numpy as np
import pandas as pd
import scipy.signal as signal
from hurst import compute_Hc
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy import stats
from scipy.fft import fft
from scipy.signal import coherence, find_peaks
from scipy.spatial import cKDTree
from scipy.special import digamma
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler
from statsmodels.graphics.tsaplots import plot_acf
from statsmodels.tsa.stattools import adfuller, grangercausalitytests
from statsmodels.tsa.vector_ar.var_model import VAR
from tqdm import tqdm


class RunLog:
    """Мини-лог для сообщений пайплайна (без зависимости от внешних классов).

    Нужен, чтобы UI/отчёты могли аккуратно показывать, что было сделано,
    и чтобы вызовы вида self.log.add(...) не падали.
    """

    def __init__(self) -> None:
        self.items: List[str] = []

    def add(self, msg: object) -> None:
        try:
            self.items.append(str(msg))
        except Exception:
            pass

    def as_text(self) -> str:
        return "\n".join(self.items)

# --- Optional dependency: nolds ---
# nolds==0.5.2 imports pkg_resources, which is removed in setuptools>=81.
# On fresh Python 3.13 installs this often breaks with:
#   ModuleNotFoundError: No module named 'pkg_resources'
# We keep the tool usable (other metrics still work) and provide a clear message.
_NOLDS_IMPORT_ERROR: Optional[str] = None
try:
    import nolds  # type: ignore
except Exception as _e:  # pragma: no cover
    nolds = None  # type: ignore
    _NOLDS_IMPORT_ERROR = str(_e)

_NOLDS_WARNED = False


def _nolds_or_warn() -> bool:
    """Return True if nolds is available; otherwise log a clear hint once."""
    global _NOLDS_WARNED
    if nolds is not None:
        return True
    if not _NOLDS_WARNED:
        _NOLDS_WARNED = True
        msg = (
            "[nolds] недоступен. Метрики, зависящие от nolds (sampen/dfa/hurst_rs fallback), "
            "будут возвращать NaN.\n"
            "Частая причина на Python 3.13: setuptools>=81 удалил pkg_resources. "
            "Решение: установи setuptools<81 (см. requirements.txt)."
        )
        if _NOLDS_IMPORT_ERROR:
            msg += f"\nПричина импорта: {_NOLDS_IMPORT_ERROR}"
        logging.error(msg)
    return False

# Импорты из нашей новой структуры
from ..config import *
from ..analysis import stats as analysis_stats
from ..metrics import connectivity
from ..metrics.registry import METRICS_REGISTRY, get_metric_func, register_metric
from ..reporting import ExcelReportWriter, HTMLReportGenerator
from ..visualization import plots

# Reuse pyplot from visualization module to keep plotting backend centralized.
plt = plots.plt

from .data_loader import load_or_generate, preprocess_timeseries
from .preprocessing import configure_warnings


# Используем константы из config.py
# Сохраняем только те, которых нет в config
# Функции загрузки данных перенесены в data_loader.py
# Используем load_or_generate из импортов

##############################################
# Функции-метрики
##############################################
def correlation_matrix(data: pd.DataFrame, **kwargs) -> np.ndarray:
    return data.corr().values

def partial_correlation_matrix(df: pd.DataFrame, control: list = None, **kwargs) -> np.ndarray:
    cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    n = len(cols)
    out = np.eye(n)
    for i in range(n):
        for j in range(i + 1, n):
            xi, xj = cols[i], cols[j]
            ctrl_vars = control if control is not None else [c for c in cols if c not in (xi, xj)]
            sub_cols = [xi, xj] + [c for c in ctrl_vars if c in cols and c not in (xi, xj)]
            sub = df[sub_cols].dropna()
            if sub.shape[0] < len(sub_cols) + 1:
                pcor = np.nan
            else:
                try:
                    R = sub.corr().values
                    P = np.linalg.pinv(R)
                    pcor = -P[0, 1] / np.sqrt(P[0, 0] * P[1, 1])
                except Exception:
                    pcor = np.nan
            out[i, j] = out[j, i] = pcor
    return out


def partial_h2_matrix(df: pd.DataFrame, control: list = None, **kwargs) -> np.ndarray:
    """Возвращает квадрат частичной корреляции (H^2) для заданного контроля."""
    pcor = partial_correlation_matrix(df, control=control, **kwargs)
    return pcor**2

def lagged_directed_correlation(df: pd.DataFrame, lag: int, **kwargs) -> np.ndarray:
    """Directed lagged correlation: M[src, tgt] = corr(src(t), tgt(t+lag)).

    Логический фикс: используем shift и совместный dropna, чтобы не рассинхронизировать время.
    """
    lag = int(max(1, lag))
    cols = list(df.columns)
    m = len(cols)
    out = np.full((m, m), np.nan, dtype=float)
    np.fill_diagonal(out, 0.0)

    for i, src in enumerate(cols):
        x = df[src]
        for j, tgt in enumerate(cols):
            if i == j:
                continue
            y = df[tgt].shift(-lag)  # y(t+lag) выравниваем с x(t)
            pair = pd.concat([x, y], axis=1).dropna()
            if pair.shape[0] <= 3:
                continue
            xv = pair.iloc[:, 0].values
            yv = pair.iloc[:, 1].values
            if np.nanstd(xv) == 0 or np.nanstd(yv) == 0:
                out[i, j] = np.nan
            else:
                out[i, j] = float(np.corrcoef(xv, yv)[0, 1])
    return out

def h2_matrix(df: pd.DataFrame, **kwargs) -> np.ndarray: return correlation_matrix(df)**2
def lagged_directed_h2(df: pd.DataFrame, lag: int, **kwargs) -> np.ndarray: return lagged_directed_correlation(df, lag)**2

def coherence_matrix(data: pd.DataFrame, **kwargs):
    """Средняя когерентность между всеми парами.

    ВАЖНО: ряды должны быть синхронизированы по общему индексу.
    Нельзя делать dropna() по каждому столбцу отдельно — иначе получишь "когерентность"
    между разными моментами времени при разнесённых NaN.
    """
    fs = float(kwargs.get("fs", 1.0))
    fs = fs if np.isfinite(fs) and fs > 0 else 1.0
    N = data.shape[1]
    coh = np.full((N, N), np.nan, dtype=float)
    np.fill_diagonal(coh, 1.0)

    for i in range(N):
        for j in range(i + 1, N):
            pair = data.iloc[:, [i, j]].dropna()
            if pair.shape[0] <= 3:
                continue

            s1 = _as_float64_1d(pair.iloc[:, 0].values)
            s2 = _as_float64_1d(pair.iloc[:, 1].values)
            n = int(min(s1.size, s2.size))
            if n <= 3:
                continue
            s1 = s1[:n]
            s2 = s2[:n]

            try:
                nperseg = int(max(8, min(64, n // 2)))
                _, Cxy = signal.coherence(s1, s2, fs=fs, nperseg=nperseg, detrend="constant")
                if Cxy.size == 0:
                    continue
                Cxy = np.clip(np.asarray(Cxy, dtype=np.float64), 0.0, 1.0)
                Cxy[~np.isfinite(Cxy)] = np.nan
                coh[i, j] = coh[j, i] = float(np.nanmean(Cxy)) if np.isfinite(Cxy).any() else np.nan
            except Exception:
                coh[i, j] = coh[j, i] = np.nan

    return coh

def _knn_entropy(X, k=DEFAULT_K_MI):
    """Вычисляет энтропию для 1D-массива с помощью KNN."""
    N = len(X)
    if N <= k: return 0.0
    tree = cKDTree(X.reshape(-1, 1))
    d, _ = tree.query(X.reshape(-1, 1), k=k + 1, p=np.inf)
    # Расстояние до k-го соседа
    r = d[:, k]
    # digamma(N) - digamma(k) + d*log(2*r_k) - это для d-мерного пространства
    # Для 1D: digamma(N) - digamma(k) + E[log(2r_k)]
    return digamma(N) - digamma(k) + np.mean(np.log(2 * r + 1e-10)) 


def _knn_mutual_info(X, Y, k=DEFAULT_K_MI):
    """KSG-оценка взаимной информации I(X;Y) через kNN (max-норма).

    Правки относительно наивных реализаций:
    - строгий eps (nextafter) для устойчивости к ties;
    - исключаем саму точку из подсчёта соседей;
    - используем ψ(nx+1), ψ(ny+1).
    """
    X = np.asarray(X, dtype=np.float64).ravel()
    Y = np.asarray(Y, dtype=np.float64).ravel()
    N = int(min(X.size, Y.size))
    if N <= k or N <= 3:
        return 0.0
    X = X[:N]
    Y = Y[:N]

    XY = np.c_[X, Y]
    tree_XY = cKDTree(XY)
    d, _ = tree_XY.query(XY, k=int(k) + 1, p=np.inf)
    eps = d[:, int(k)]
    # строгий радиус
    eps = np.nextafter(eps, 0.0)

    tree_X = cKDTree(X.reshape(-1, 1))
    tree_Y = cKDTree(Y.reshape(-1, 1))

    nx = np.fromiter(
        (max(0, len(tree_X.query_ball_point([X[i]], r=float(eps[i]), p=np.inf)) - 1) for i in range(N)),
        dtype=float,
        count=N,
    )
    ny = np.fromiter(
        (max(0, len(tree_Y.query_ball_point([Y[i]], r=float(eps[i]), p=np.inf)) - 1) for i in range(N)),
        dtype=float,
        count=N,
    )

    raw_mi = digamma(N) + digamma(int(k)) - np.mean(digamma(nx + 1.0) + digamma(ny + 1.0))
    if not np.isfinite(raw_mi):
        return float("nan")
    return float(max(0.0, raw_mi))

def mutual_info_matrix(data: pd.DataFrame, k=DEFAULT_K_MI, **kwargs):
    """Матрица взаимной информации (KSG kNN).

    Важно: используем совместный dropna по паре, чтобы не рассинхронизировать время.
    """
    cols = list(data.columns)
    n_vars = len(cols)
    mi = np.zeros((n_vars, n_vars), dtype=float)

    for i in range(n_vars):
        for j in range(i + 1, n_vars):
            pair = data.iloc[:, [i, j]].dropna()
            if pair.shape[0] <= k:
                mi[i, j] = mi[j, i] = np.nan
                continue
            s1 = pair.iloc[:, 0].values
            s2 = pair.iloc[:, 1].values
            val = _knn_mutual_info(s1, s2, k=k)
            mi[i, j] = mi[j, i] = val

    return mi

def _knn_conditional_mutual_info(X, Y, Z, k=DEFAULT_K_MI):
    """KSG-оценка условной взаимной информации I(X;Y|Z) через kNN (max-норма).

    Важно:
    - строгий eps (nextafter) для ties;
    - исключаем саму точку;
    - используем ψ(n+1).
    """
    X = np.asarray(X, dtype=np.float64).ravel()
    Y = np.asarray(Y, dtype=np.float64).ravel()
    Z = np.asarray(Z, dtype=np.float64)
    if Z.ndim == 1:
        Z = Z.reshape(-1, 1)
    N = int(min(X.size, Y.size, Z.shape[0]))
    if N <= k or N <= 3:
        return 0.0
    X = X[:N]
    Y = Y[:N]
    Z = Z[:N, :]

    XZ = np.c_[X, Z]
    YZ = np.c_[Y, Z]
    XYZ = np.c_[X, Y, Z]

    tree_XYZ = cKDTree(XYZ)
    d, _ = tree_XYZ.query(XYZ, k=int(k) + 1, p=np.inf)
    eps = d[:, int(k)]
    eps = np.nextafter(eps, 0.0)

    tree_XZ = cKDTree(XZ)
    tree_YZ = cKDTree(YZ)
    tree_Z = cKDTree(Z)

    nxz = np.fromiter(
        (max(0, len(tree_XZ.query_ball_point(XZ[i], r=float(eps[i]), p=np.inf)) - 1) for i in range(N)),
        dtype=float,
        count=N,
    )
    nyz = np.fromiter(
        (max(0, len(tree_YZ.query_ball_point(YZ[i], r=float(eps[i]), p=np.inf)) - 1) for i in range(N)),
        dtype=float,
        count=N,
    )
    nz = np.fromiter(
        (max(0, len(tree_Z.query_ball_point(Z[i], r=float(eps[i]), p=np.inf)) - 1) for i in range(N)),
        dtype=float,
        count=N,
    )

    cmi = digamma(int(k)) - np.mean(digamma(nxz + 1.0) + digamma(nyz + 1.0) - digamma(nz + 1.0))
    if not np.isfinite(cmi):
        return float("nan")
    return float(max(0.0, cmi))

def mutual_info_matrix_partial(
    data: pd.DataFrame,
    control: Optional[List[str]] = None,
    k=DEFAULT_K_MI,
    **kwargs,
):
    """Матрица условной взаимной информации I(X;Y|Z) (KSG kNN).

    Если control=None: Z = все остальные переменные.
    Важно: формируем sub-таблицу и делаем dropna ОДИН РАЗ (иначе рассинхрон).
    """
    cols = list(data.columns)
    N = len(cols)
    pmi = np.zeros((N, N), dtype=float)

    for i in range(N):
        for j in range(i + 1, N):
            xi, xj = cols[i], cols[j]
            Z_cols = control if control is not None else [c for c in cols if c not in (xi, xj)]
            Z_cols = [c for c in Z_cols if c in data.columns and c not in (xi, xj)]

            if not Z_cols:
                pair = data[[xi, xj]].dropna()
                if pair.shape[0] <= k:
                    val = np.nan
                else:
                    val = _knn_mutual_info(pair[xi].values, pair[xj].values, k=k)
            else:
                sub = data[[xi, xj] + Z_cols].dropna()
                if sub.shape[0] <= k:
                    val = np.nan
                else:
                    X = sub[xi].values
                    Y = sub[xj].values
                    Z = sub[Z_cols].values
                    val = _knn_conditional_mutual_info(X, Y, Z, k=k)

            pmi[i, j] = pmi[j, i] = val

    return pmi

def compute_granger_matrix(df: pd.DataFrame, lags: int = DEFAULT_MAX_LAG, **kwargs) -> np.ndarray:
    n = df.shape[1]
    G = np.full((n, n), 1.0)
    cols = df.columns.tolist()
    # matrix[src, tgt] = pvalue(src -> tgt)
    for src in range(n):
        for tgt in range(n):
            if src == tgt:
                G[src, tgt] = 0.0
                continue
            data_pair = df[[cols[tgt], cols[src]]].dropna()  # [target, source]
            if len(data_pair) > lags * 2 + 5:
                try:
                    tests = grangercausalitytests(data_pair, maxlag=lags)
                    G[src, tgt] = tests[lags][0]['ssr_ftest'][1]
                except (np.linalg.LinAlgError, ValueError):
                    G[src, tgt] = np.nan
    return G

def _load_pyinform():
    """Ленивая загрузка pyinform без падения всего приложения."""
    if not PYINFORM_AVAILABLE:
        logging.warning(
            "[PyInform] pyinform не установлен: TE будет посчитан через fallback (дискретизация + эмпирические вероятности).",
        )
        return None
    return importlib.import_module("pyinform")



def _transfer_entropy_discrete(source_d: np.ndarray, target_d: np.ndarray, k: int = 1) -> float:
    """Эмпирическая transfer entropy для дискретных целочисленных рядов.

    Конвенция: TE(source -> target).
    Реализация: sum p(x_{t}|x_{t-k:t-1}, y_{t-k:t-1}) * log( p(x_t|x_past,y_past) / p(x_t|x_past) )
    """
    try:
        k = int(k)
        if k < 1:
            k = 1
        source_d = np.asarray(source_d, dtype=int).ravel()
        target_d = np.asarray(target_d, dtype=int).ravel()
        n = min(source_d.size, target_d.size)
        if n <= k + 1:
            return float("nan")
        source_d = source_d[:n]
        target_d = target_d[:n]
        n_eff = n - k

        c_xyz = Counter()
        c_xx = Counter()
        c_xpast_ypast = Counter()
        c_xpast = Counter()

        for t in range(k, n):
            x_next = int(target_d[t])
            x_past = tuple(int(v) for v in target_d[t - k : t])
            y_past = tuple(int(v) for v in source_d[t - k : t])

            c_xyz[(x_next, x_past, y_past)] += 1
            c_xx[(x_next, x_past)] += 1
            c_xpast_ypast[(x_past, y_past)] += 1
            c_xpast[x_past] += 1

        te = 0.0
        for (x_next, x_past, y_past), c in c_xyz.items():
            denom_xy = c_xpast_ypast.get((x_past, y_past), 0)
            denom_x = c_xpast.get(x_past, 0)
            if denom_xy <= 0 or denom_x <= 0:
                continue
            p1 = c / denom_xy  # p(x_next | x_past, y_past)
            p0 = c_xx.get((x_next, x_past), 0) / denom_x  # p(x_next | x_past)
            if p0 <= 0 or p1 <= 0:
                continue
            te += (c / n_eff) * float(np.log(p1 / p0))

        # численная защита
        if not np.isfinite(te):
            return float("nan")
        return float(max(0.0, te))
    except Exception:
        return float("nan")


def compute_TE(source: np.ndarray, target: np.ndarray, lag: int = 1, bins: int = DEFAULT_BINS):
    """Transfer Entropy (дискретная).

    Семантика параметра lag: длина истории k (как в pyinform.transfer_entropy k=...).

    Логические фиксы:
    - дискретизация через квантильные бины (устойчивее, чем min/max линейка);
    - перед бинингом z-score (масштаб-инвариантность);
    - маленький детерминированный jitter при ties (квантизация/повторы значений).
    """
    try:
        def _zscore_1d(x: np.ndarray) -> np.ndarray:
            x = np.asarray(x, dtype=np.float64).ravel()
            if x.size == 0:
                return x
            m = np.nanmean(x)
            s = np.nanstd(x)
            if not np.isfinite(s) or s <= 0:
                return x - m
            return (x - m) / s

        def _add_tiny_jitter(x: np.ndarray) -> np.ndarray:
            # детерминированный jitter для борьбы с ties (важно для kNN/квантилей)
            x = np.asarray(x, dtype=np.float64)
            if x.size <= 3:
                return x
            # если много повторов — чуть шевелим
            uniq = np.unique(x[np.isfinite(x)])
            if uniq.size < max(3, int(0.2 * x.size)):
                rng = np.random.default_rng(0)
                scale = (np.nanstd(x) if np.nanstd(x) > 0 else 1.0) * 1e-10
                x = x + rng.normal(0.0, scale, size=x.shape)
            return x

        def discretize_quantile(series: np.ndarray, num_bins: int) -> np.ndarray:
            s = np.asarray(series, dtype=np.float64).ravel()
            if s.size == 0:
                return np.array([], dtype=int)
            s = _add_tiny_jitter(_zscore_1d(s))
            if not np.isfinite(s).all():
                s = s[np.isfinite(s)]
            if s.size == 0:
                return np.array([], dtype=int)
            if float(np.nanmin(s)) == float(np.nanmax(s)):
                return np.zeros(int(series.size), dtype=int)

            q = np.linspace(0.0, 1.0, int(num_bins) + 1)
            edges = np.quantile(s, q)
            # убираем дубли рёбер (если данных мало/много ties)
            edges = np.unique(edges)
            if edges.size <= 2:
                return np.zeros(int(series.size), dtype=int)

            # делаем края "открытыми" справа
            edges[-1] = np.nextafter(edges[-1], edges[-1] + 1.0)
            disc = np.digitize(_add_tiny_jitter(_zscore_1d(np.asarray(series))), bins=edges[1:-1], right=False)
            disc = np.clip(disc, 0, int(num_bins) - 1)
            return disc.astype(int)

        source_discrete = discretize_quantile(source, bins)
        target_discrete = discretize_quantile(target, bins)

        pyinform = _load_pyinform()
        k = int(max(1, lag))
        if pyinform is not None:
            return float(pyinform.transfer_entropy(source_discrete, target_discrete, k=k))

        return _transfer_entropy_discrete(source_discrete, target_discrete, k=k)

    except Exception as e:
        logging.error(f"[TE] Ошибка вычисления: {e}")
        return float("nan")

def TE_matrix(df: pd.DataFrame, lag: int = 1, bins: int = DEFAULT_BINS, **kwargs):
    """
    Строит матрицу Transfer Entropy для всех пар.

    Конвенция для направленных матриц: M[src, tgt] = мера src → tgt.
    """
    n = df.shape[1]
    te_matrix = np.zeros((n, n))

    for src in range(n):
        for tgt in range(n):
            if src == tgt:
                continue

            pair = df.iloc[:, [src, tgt]].dropna()
            if len(pair) <= lag:
                te_matrix[src, tgt] = np.nan
                continue
            s_src = pair.iloc[:, 0].values
            s_tgt = pair.iloc[:, 1].values
            te_matrix[src, tgt] = compute_TE(s_src, s_tgt, lag=lag, bins=bins)

    return te_matrix

def TE_matrix_partial(
    df: pd.DataFrame,
    lag: int = 1,
    control: Optional[List[str]] = None,
    bins: int = DEFAULT_BINS,
) -> np.ndarray:
    """
    Приближённая "partial" Transfer Entropy.

    Практичная аппроксимация:
      1) линейно вычитаем влияние контрольных переменных из src и tgt (остатки OLS);
      2) считаем обычный TE между остатками.

    Конвенция: M[src, tgt] = мера src → tgt.
    """
    cols = list(df.columns)
    N = len(cols)
    M = np.zeros((N, N))

    def _residualize(y: np.ndarray, X: np.ndarray) -> np.ndarray:
        if X.size == 0:
            return y
        X_aug = np.c_[np.ones(len(X)), X]
        beta, *_ = np.linalg.lstsq(X_aug, y, rcond=None)
        return y - X_aug @ beta

    for i, src in enumerate(cols):
        for j, tgt in enumerate(cols):
            if i == j:
                continue

            ctrl_cols = control if control is not None else [c for c in cols if c not in (src, tgt)]
            ctrl_cols = [c for c in ctrl_cols if c in df.columns]

            use_cols = [src, tgt] + ctrl_cols
            sub = df[use_cols].dropna()
            if sub.shape[0] <= lag + 1:
                M[i, j] = np.nan
                continue

            s_src = sub[src].values
            s_tgt = sub[tgt].values
            X_ctrl = sub[ctrl_cols].values if ctrl_cols else np.empty((len(sub), 0))

            try:
                s_src_r = _residualize(s_src, X_ctrl)
                s_tgt_r = _residualize(s_tgt, X_ctrl)
                M[i, j] = compute_TE(s_src_r, s_tgt_r, lag=lag, bins=bins)
            except Exception:
                M[i, j] = np.nan

    return M


def AH_matrix(df: pd.DataFrame, embed_dim=DEFAULT_EMBED_DIM, tau=DEFAULT_EMBED_TAU) -> np.ndarray:
    """
    Конвенция: M[src, tgt] = мера src → tgt.
    """
    df = df.dropna(axis=0, how='any')
    N = df.shape[1]
    out = np.zeros((N, N))
    arr = df.values
    for src in range(N):
        for tgt in range(N):
            if src == tgt:
                out[src, tgt] = 0.0
                continue

            # _H_ratio_direction(X, Y): интерпретируется как X → Y
            H_val = _H_ratio_direction(arr[:, src], arr[:, tgt], m=embed_dim, tau=tau)
            if H_val is None or H_val <= 0:
                AH = 0.0
            else:
                AH = 1.0 / H_val
                if AH > 1.0:
                    AH = 1.0
            out[src, tgt] = AH
    return out

def _H_ratio_direction(X, Y, m=DEFAULT_EMBED_DIM, tau=DEFAULT_EMBED_TAU):
    n = len(X)
    if len(Y) != n or n < 2:
        return None
    L = n - (m - 1) * tau
    if L < 2:
        return None
    
    X_state = np.zeros((L, m))
    Y_state = np.zeros((L, m))
    for j in range(m):
        X_state[:, j] = X[j*tau : j*tau + L]
        Y_state[:, j] = Y[j*tau : j*tau + L]
    
    valid_indices = ~np.isnan(X_state).any(axis=1) & ~np.isnan(Y_state).any(axis=1)
    if not np.any(valid_indices):
        return None
        
    X_state_valid = X_state[valid_indices]
    Y_state_valid = Y_state[valid_indices]
    
    if len(X_state_valid) < 2:
        return None

    tree_X = cKDTree(X_state_valid)
    tree_Y = cKDTree(Y_state_valid)
    
    dists_X, idx_X = tree_X.query(X_state_valid, k=2)
    
    
    if idx_X.shape[1] < 2: 
        return None 
    
    nn_idx = idx_X[:, 1] 
    diff = Y_state_valid - Y_state_valid[nn_idx]
    dY1 = np.sqrt(np.sum(diff**2, axis=1))
    
    dists_Y, _ = tree_Y.query(Y_state_valid, k=2)
    dY2 = dists_Y[:, 1]
    dY2 = np.where(dY2 == 0, 1e-10, dY2) 
    
    ratios = dY1 / dY2
    ratios = ratios[np.isfinite(ratios)] 
    
    if len(ratios) == 0:
        return None

    H_val = np.mean(ratios)
    return H_val

def compute_partial_AH_matrix(data: pd.DataFrame,
                               max_lag: int = DEFAULT_MAX_LAG,
                               embed_dim: int = DEFAULT_EMBED_DIM,
                               tau: int = DEFAULT_EMBED_TAU,
                               control: List[str] = None) -> np.ndarray:
    df = data.dropna(axis=0, how='any')
    N = df.shape[1]
    if N < 2:
        return np.zeros((N, N))

    if control and len(control) > 0:
        resid_df = pd.DataFrame(index=df.index)
        for col in df.columns:
            X_ctrl = df[control]
            y = df[col]
            if len(X_ctrl) > 0 and len(y) == len(X_ctrl) and not X_ctrl.isnull().any().any():
                try:
                    model = LinearRegression().fit(X_ctrl.values, y.values)
                    resid = y.values - model.predict(X_ctrl.values)
                    resid_df[col] = resid
                except ValueError: 
                    resid_df[col] = y 
            else:
                resid_df[col] = y
    else:
        try:
            model = VAR(df.values)
            res_full = model.fit(max_lag, ic=None)
            resid_df = pd.DataFrame(res_full.resid, columns=df.columns)
        except Exception as e:
            logging.error(f"VAR fit error (partial AH, fallback to raw): {e}")
            resid_df = df 

    return AH_matrix(resid_df, embed_dim=embed_dim, tau=tau)


def directional_AH_matrix(df: pd.DataFrame, maxlags: int = 5) -> np.ndarray:
    return AH_matrix(df, embed_dim=DEFAULT_EMBED_DIM, tau=DEFAULT_EMBED_TAU)

def granger_dict(df: pd.DataFrame, maxlag: int = 4) -> dict:
    results = {}
    cols = list(df.columns)
    for i, tgt in enumerate(cols):
        for j, src in enumerate(cols):
            if i == j:
                continue
            sub = df[[tgt, src]].dropna()
            if len(sub) < (maxlag + 10):
                results[f"{src}->{tgt}"] = None
                continue
            try:
                tests = grangercausalitytests(sub, maxlag=maxlag)
            except Exception as e:
                logging.error(f"[Granger] Ошибка Granger для {src}->{tgt}: {e}")
                results[f"{src}->{tgt}"] = None
                continue 
            results[f"{src}->{tgt}"] = tests # Сохраняем РЕЗУЬТАТ
    return results

# эта матрица НЕ ТА ЖЕ ЧТО В МАППИНГЕ
def _compute_granger_matrix_internal(df: pd.DataFrame, lags: int = DEFAULT_MAX_LAG) -> np.ndarray:
    n = df.shape[1]
    G = np.zeros((n, n))
    cols = df.columns.tolist()
    # matrix[src, tgt] = pvalue(src -> tgt)
    for src in range(n):
        for tgt in range(n):
            if src == tgt:
                G[src, tgt] = 0.0
                continue
            sub = df[[cols[tgt], cols[src]]].dropna()  # [target, source]
            try:
                tests = grangercausalitytests(sub, maxlag=lags)
                pvals = [tests[l][0]['ssr_ftest'][1] for l in tests]
                G[src, tgt] = min(pvals)
            except Exception as e:
                logging.error(f"[Granger-Internal] Ошибка Granger для {cols[src]}->{cols[tgt]}: {e}")
                G[src, tgt] = np.nan
    return G


def compute_partial_granger_matrix(data: pd.DataFrame, lags=DEFAULT_MAX_LAG) -> np.ndarray:
    """
    контроль остальных переменных, грейндж
    """
    df = data.dropna(axis=0, how='any')
    N = df.shape[1]
    if N < 3:
        return _compute_granger_matrix_internal(data, lags=lags)
    pg_matrix = np.zeros((N, N))
    T = len(df)
    p = lags
    if T <= p:
        return pg_matrix
    arr = df.values
    try:
        model_full = VAR(arr)
        res_full = model_full.fit(p, ic=None)
    except Exception as e:
        logging.error(f"VAR fit error (partial Granger): {e}")
        return pg_matrix
    sigma_full = np.cov(res_full.resid, rowvar=False)
    for i in range(N):
        reduced_arr = np.delete(arr, i, axis=1)
        try:
            model_red = VAR(reduced_arr)
            res_red = model_red.fit(p, ic=None)
            sigma_red = np.cov(res_red.resid, rowvar=False)
        except Exception as e:
            for j in range(N):
                if j != i:
                    pg_matrix[i, j] = np.nan 
            continue
        for j in range(N):
            if i == j:
                pg_matrix[i, j] = 0.0
            else:
                idx_j = j - 1 if i < j else j
                var_full = sigma_full[j, j] if sigma_full.shape[0] > j else np.var(res_full.resid[:, j])
                var_red = sigma_red[idx_j, idx_j] if sigma_red.shape[0] > idx_j else np.var(res_red.resid[:, idx_j])
                if var_full <= 0 or var_red <= 0:
                    gc_val = np.nan 
                else:
                    gc_val = np.log(var_red / var_full)
                    if gc_val < 0:
                        gc_val = 0.0
                pg_matrix[i, j] = gc_val
    return pg_matrix

def p_to_score(p: float, eps: float = 1e-300) -> float:
    """Convert p-value to comparable strength score: -log10(p). Higher = stronger."""
    if p is None or (isinstance(p, float) and np.isnan(p)):
        return np.nan
    p = float(p)
    if p <= 0:
        p = eps
    return float(-np.log10(p))


def granger_matrix(df: pd.DataFrame, granger_dict_result: dict) -> np.ndarray:
    cols = list(df.columns)
    n_vars = len(cols)
    G = np.ones((n_vars, n_vars))
    for i, tgt in enumerate(cols):
        for j, src in enumerate(cols):
            if i == j:
                G[i, j] = 0
            else:
                key = f"{src}->{tgt}"
                if granger_dict_result.get(key) is None:
                    G[i, j] = np.nan
                else:
                    test_dict = granger_dict_result[key]
                    bp = 1.0 
                    found_valid_p = False 
                    for lag_val, dct in test_dict.items():
                        if isinstance(dct, list) and len(dct) > 0 and 'ssr_ftest' in dct[0]:
                            F, pval, _, _ = dct[0]['ssr_ftest']
                            if not np.isnan(pval): 
                                bp = min(bp, pval)
                                found_valid_p = True
                    G[i, j] = bp if found_valid_p else np.nan 
    return G



def remove_linear_dependency(sub: pd.DataFrame, src: str, tgt: str, control_cols: list) -> tuple[np.ndarray, np.ndarray]:
    """Удаляет линейную компоненту контролей из src/tgt и возвращает остатки (src_res, tgt_res)."""
    if not control_cols:
        r1 = sub[src].to_numpy(dtype=float)
        r2 = sub[tgt].to_numpy(dtype=float)
        return r1, r2

    X = sub[control_cols].to_numpy(dtype=float)
    y_src = sub[src].to_numpy(dtype=float)
    y_tgt = sub[tgt].to_numpy(dtype=float)

    mdl_src = LinearRegression().fit(X, y_src)
    mdl_tgt = LinearRegression().fit(X, y_tgt)

    r1 = y_src - mdl_src.predict(X)
    r2 = y_tgt - mdl_tgt.predict(X)
    return r1, r2

def granger_matrix_partial(df: pd.DataFrame, maxlag=DEFAULT_MAX_LAG, control=None) -> np.ndarray:
    """Conditional Granger causality p-values via multivariate VAR.

    Конвенция: G[src, tgt] = pvalue( src -> tgt ), т.е. тест "src НЕ вызывает tgt".
    control: список контролей (без src/tgt). Если None — все остальные переменные.
    """
    columns = list(df.columns)
    n = len(columns)
    G = np.full((n, n), np.nan, dtype=float)
    np.fill_diagonal(G, 0.0)

    for src_i, src in enumerate(columns):
        for tgt_j, tgt in enumerate(columns):
            if src_i == tgt_j:
                continue

            control_cols = control if control is not None else [c for c in columns if c not in (src, tgt)]
            control_cols = [c for c in control_cols if c in df.columns and c not in (src, tgt)]
            use_cols = [tgt, src] + control_cols  # порядок важен только для удобства
            sub = df[use_cols].dropna()

            # минимум наблюдений: грубо (кол-во параметров VAR растёт с p и k)
            p = int(max(1, maxlag))
            k_vars = len(use_cols)
            if sub.shape[0] < max(30, 5 * p * k_vars):
                continue

            try:
                res = VAR(sub).fit(maxlags=p, ic=None, trend="c")
                test = res.test_causality(caused=tgt, causing=[src], kind="f")
                G[src_i, tgt_j] = float(test.pvalue) if np.isfinite(test.pvalue) else np.nan
            except Exception:
                G[src_i, tgt_j] = np.nan

    return G

def plt_fft_analysis(series: pd.Series, *, fs: float = 1.0):
    """Быстрый FFT-анализ.

    Важно: частотная шкала зависит от частоты дискретизации fs (Гц).
    По умолчанию fs=1.0 (частоты в "циклах на отсчёт").
    """
    arr = _as_float64_1d(series.dropna().values)
    if arr.size == 0:
        return np.array([]), np.array([]), np.array([]), np.array([])
    n = int(arr.size)
    fs = float(fs) if fs and np.isfinite(fs) and fs > 0 else 1.0
    dt = 1.0 / fs
    freqs = np.fft.fftfreq(n, d=dt)
    fft_vals = fft(arr)
    amplitude = np.abs(fft_vals)
    phase = np.angle(fft_vals)
    pos_mask = freqs >= 0
    freqs, amplitude, phase = freqs[pos_mask], amplitude[pos_mask], phase[pos_mask]
    peaks, _ = find_peaks(amplitude, height=(np.max(amplitude) * 0.2 if amplitude.size > 0 else 0))
    logging.debug(f"[FFT] Найдено пиков на частотах: {freqs[peaks] if peaks.size > 0 else 'Нет пиков'}")
    return freqs, amplitude, phase, peaks

def plot_amplitude_response(series: pd.Series, title: str) -> BytesIO:
    freqs, amplitude, phase, peaks = plt_fft_analysis(series)
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.plot(freqs, amplitude, label="АЧХ")
    if peaks.size > 0:
        ax.plot(freqs[peaks], amplitude[peaks], "x", label="Пики")
    ax.set_title(title)
    ax.set_xlabel("Частота")
    ax.set_ylabel("Амплитуда")
    ax.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=100)
    buf.seek(0)
    plt.close(fig)
    return buf

def plot_phase_response(series: pd.Series, title: str) -> BytesIO:
    freqs, amplitude, phase, peaks = plt_fft_analysis(series)
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.plot(freqs, phase, label="ФЧХ", color="orange")
    ax.set_title(title)
    ax.set_xlabel("Частота")
    ax.set_ylabel("Фаза (рад)")
    ax.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=100)
    buf.seek(0)
    plt.close(fig)
    return buf

def plot_combined_ac_fch(data: pd.DataFrame, title: str) -> BytesIO:
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8, 8), sharex=True)
    for col in data.columns:
        series = data[col]
        freqs, amplitude, phase, _ = plt_fft_analysis(series)
        if freqs.size > 0:
            ax1.plot(freqs, amplitude, label=col)
            ax2.plot(freqs, phase, label=col)
    ax1.set_title(title + " - АЧХ")
    ax1.set_ylabel("Амплитуда")
    ax1.legend()
    ax2.set_title(title + " - ФЧХ")
    ax2.set_xlabel("Частота")
    ax2.set_ylabel("Фаза (рад)")
    ax2.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=100)
    buf.seek(0)
    plt.close(fig)
    return buf

##############################################
# индивидуальный AC & PH для каждого ряда
##############################################
def plot_individual_ac_ph(data: pd.DataFrame, title: str) -> dict:
    plots = {}
    for col in data.columns:
        series = data[col]
        # График АЧХ
        fig1, ax1 = plt.subplots(figsize=(6, 4))
        freqs, amplitude, _, peaks = plt_fft_analysis(series)
        ax1.plot(freqs, amplitude, label=f"АЧХ {col}")
        if peaks.size > 0:
            ax1.plot(freqs[peaks], amplitude[peaks], "x", label="Пики")
        ax1.set_title(f"АЧХ {col}")
        ax1.set_xlabel("Частота")
        ax1.set_ylabel("Амплитуда")
        ax1.legend()
        buf1 = BytesIO()
        plt.tight_layout()
        plt.savefig(buf1, format="png", dpi=100)
        buf1.seek(0)
        plt.close(fig1)
        # График ФЧХ
        fig2, ax2 = plt.subplots(figsize=(6, 4))
        freqs, _, phase, _ = plt_fft_analysis(series)
        ax2.plot(freqs, phase, label=f"ФЧХ {col}", color="orange")
        ax2.set_title(f"ФЧХ {col}")
        ax2.set_xlabel("Частота")
        ax2.set_ylabel("Фаза (рад)")
        ax2.legend()
        buf2 = BytesIO()
        plt.tight_layout()
        plt.savefig(buf2, format="png", dpi=100)
        buf2.seek(0)
        plt.close(fig2)
        plots[col] = {"AC": buf1, "PH": buf2}
    return plots

##############################################
# sample entropy
##############################################
def compute_sample_entropy(series: pd.Series) -> float:
    """Compatibility wrapper for sample entropy computation."""
    return analysis_stats.compute_sample_entropy(series)


def _as_float64_1d(x) -> np.ndarray:
    """Безопасно приводит вход к 1D float64 без NaN/inf."""
    arr = np.asarray(x, dtype=np.float64).reshape(-1)
    if arr.size == 0:
        return arr
    m = np.isfinite(arr)
    return arr[m]

##############################################
# Функции для частотного анализа когерентности
##############################################
def plot_coherence_vs_frequency(
    series1: pd.Series,
    series2: pd.Series,
    title: str,
    *,
    fs: float = 1.0,
    nperseg: Optional[int] = None,
) -> BytesIO:
    s1 = _as_float64_1d(series1.dropna().values)
    s2 = _as_float64_1d(series2.dropna().values)
    n = int(min(s1.size, s2.size))
    if n <= 3:
        return BytesIO()
    s1 = s1[:n]
    s2 = s2[:n]
    fs = float(fs) if fs and np.isfinite(fs) and fs > 0 else 1.0
    if nperseg is None:
        # для коротких рядов дефолт scipy (256) даёт вырожденную оценку
        nperseg = int(max(8, min(64, n // 2)))
    nperseg = int(max(8, min(nperseg, n)))
    freqs, cxy = coherence(s1, s2, fs=fs, nperseg=nperseg, detrend="constant")
    if cxy.size:
        cxy = np.clip(np.asarray(cxy, dtype=np.float64), 0.0, 1.0)
        cxy[~np.isfinite(cxy)] = np.nan
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.plot(freqs, cxy, label="Когерентность")
    if cxy.size > 0:
        # np.nanargmax падает на all-nan
        max_idx = int(np.nanargmax(cxy)) if np.isfinite(cxy).any() else 0
        max_freq = freqs[max_idx]
        max_coh = cxy[max_idx]
        ax.plot(max_freq, max_coh, "ro", label=f"Макс. связь: {max_coh:.3f} на {max_freq:.3f}Hz")
        ax.annotate(f"{max_freq:.3f} Hz", xy=(max_freq, max_coh), xytext=(max_freq, max_coh+0.05),
                    arrowprops=dict(facecolor='black', shrink=0.05))
    ax.set_title(title)
    ax.set_xlabel("Частота (Hz)")
    ax.set_ylabel("Когерентность")
    ax.set_ylim(0, 1)
    ax.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png")
    buf.seek(0)
    plt.close(fig)
    return buf

##############################################
# Функции для экспорта данных в Excel
##############################################
def add_raw_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Добавляет лист с исходными данными."""
    ws = wb.create_sheet("Raw Data")
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))

def plot_heatmap(matrix: np.ndarray, title: str, legend_text: str = "", annotate: bool = False, vmin=None, vmax=None) -> BytesIO:
    """Compatibility wrapper around src.visualization.plots.plot_heatmap."""
    return plots.plot_heatmap(matrix, title, legend_text=legend_text, annotate=annotate, vmin=vmin, vmax=vmax)

def plot_connectome(matrix: np.ndarray, method_name: str, threshold: float = 0.2,
                    directed: bool = False, invert_threshold: bool = False, legend_text: str = "") -> BytesIO:
    """Compatibility wrapper around src.visualization.plots.plot_connectome."""
    return plots.plot_connectome(
        matrix,
        method_name,
        threshold=threshold,
        directed=directed,
        invert_threshold=invert_threshold,
        legend_text=legend_text,
    )

def add_method_to_sheet(ws, row: int, title: str, matrix: np.ndarray, directed: bool = False, legend_text: str = "") -> int:
    ws.append([title])
    if matrix is None:
        ws.append(["Метод не работает для этих данных."])
        return ws.max_row
    df_mat = pd.DataFrame(matrix)
    for r in dataframe_to_rows(df_mat, index=False, header=True):
        ws.append(r)
    buf_heat = plot_heatmap(matrix, title + " Heatmap", legend_text=legend_text)
    img_heat = Image(buf_heat)
    img_heat.width = 400
    img_heat.height = 300
    ws.add_image(img_heat, f"A{ws.max_row + 2}")
    buf_conn = plot_connectome(matrix, title + " Connectome", threshold=0.2, directed=directed, invert_threshold=False, legend_text=legend_text)
    img_conn = Image(buf_conn)
    img_conn.width = 400
    img_conn.height = 400
    ws.add_image(img_conn, f"G{ws.max_row + 2}")
    return ws.max_row

def fmt_val(v):
    try:
        f = float(v)
        if np.isnan(f):
            return "N/A"
        return f"{f:.3f}"
    except Exception:
        return "N/A"

def fdr_bh(pvals: np.ndarray) -> np.ndarray:
    """Benjamini–Hochberg FDR correction. Returns q-values (same shape)."""
    p = np.asarray(pvals, dtype=float)
    q = np.full(p.shape, np.nan, dtype=float)
    mask = np.isfinite(p)
    if mask.sum() == 0:
        return q
    pv = p[mask].ravel()
    m = pv.size
    order = np.argsort(pv)
    ranked = pv[order]
    q_raw = ranked * m / (np.arange(1, m + 1))
    # monotone
    q_mono = np.minimum.accumulate(q_raw[::-1])[::-1]
    q_mono = np.clip(q_mono, 0.0, 1.0)
    out = np.empty_like(pv)
    out[order] = q_mono
    q[mask] = out
    return q

def apply_pvalue_correction_matrix(mat: np.ndarray, directed: bool) -> np.ndarray:
    """Apply FDR correction to a p-value matrix (off-diagonal entries)."""
    M = np.array(mat, dtype=float, copy=True)
    n = M.shape[0]
    if n == 0:
        return M
    mask = np.isfinite(M)
    # ignore diagonal
    np.fill_diagonal(mask, False)
    if not directed:
        # use only upper triangle to avoid double-counting, then mirror
        tri = np.triu(mask, 1)
        q = fdr_bh(M[tri])
        M[tri] = q
        M = M + M.T
        np.fill_diagonal(M, 0.0)
        return M
    else:
        q = fdr_bh(M[mask])
        M[mask] = q
        np.fill_diagonal(M, 0.0)
        return M

#
# МАППИНГ
#
# МАППИНГ
###
# Metric registry lives in src.metrics.registry to keep engine as orchestrator.
method_mapping = dict(METRICS_REGISTRY)

# AH methods are still implemented in this module, so they are registered here.
def _metric_ah_full(data: pd.DataFrame, lag: int = 1, control=None, **kwargs) -> np.ndarray:
    return AH_matrix(data)


def _metric_ah_partial(data: pd.DataFrame, lag: int = 1, control=None, **kwargs) -> np.ndarray:
    return compute_partial_AH_matrix(data, max_lag=lag, control=control)


def _metric_ah_directed(data: pd.DataFrame, lag: int = 1, control=None, **kwargs) -> np.ndarray:
    return AH_matrix(data) if not control else compute_partial_AH_matrix(data, max_lag=lag, control=control)


for _name, _func in {
    "ah_full": _metric_ah_full,
    "ah_partial": _metric_ah_partial,
    "ah_directed": _metric_ah_directed,
}.items():
    register_metric(_name, _func)

method_mapping = dict(METRICS_REGISTRY)


@dataclass(frozen=True)
class MethodSpec:
    directed: bool
    is_p_value: bool
    control_dependent: bool
    supports_lag: bool
    description: str = ""

# -----------------------------
# Метод-метаданные и логика
# (PVAL_METHODS, DIRECTED_METHODS, METHOD_INFO импортированы из config.py
#  через `from ..config import *` выше)
# -----------------------------

# is_pvalue_method, is_directed_method, is_control_sensitive_method
# уже импортированы из config.py через `from ..config import *`


def _pair_score(variant: str, mat: np.ndarray, i: int, j: int) -> float:
    """Скалярная метрика для одной пары (i,j) из матрицы связности.

    Используется для 3D-"кубиков" по парам.

    Правила:
    - p-value методы: чем меньше p, тем "лучше" связь -> score = 1 - p (в [0,1])
    - directed методы: берём i -> j
    - undirected: берём max(|A[i,j]|, |A[j,i]|)
    """
    if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
        return float("nan")
    n = int(mat.shape[0])
    if i < 0 or j < 0 or i >= n or j >= n or i == j:
        return float("nan")

    try:
        if is_directed_method(variant):
            v = float(mat[i, j])
        else:
            v = float(mat[i, j])
            v2 = float(mat[j, i])
            v = float(max(abs(v), abs(v2)))

        if is_pvalue_method(variant):
            # p-value: меньше = лучше -> инвертируем.
            if not np.isfinite(v):
                return float("nan")
            v = float(np.clip(v, 0.0, 1.0))
            return float(1.0 - v)

        return float(abs(v)) if np.isfinite(v) else float("nan")
    except Exception:
        return float("nan")


# is_control_sensitive_method — из config.py


# METHOD_INFO — из config.py (через `from ..config import *`)


def _is_pvalue_method(variant: str) -> bool:
    return is_pvalue_method(variant)


def _is_directed_method(variant: str) -> bool:
    return is_directed_method(variant)

def _lag_quality(variant: str, mat: np.ndarray) -> float:
    """Скалярная метрика качества лага: больше => лучше (единая конвенция)."""
    if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
        return np.nan
    n = mat.shape[0]
    if n < 2:
        return np.nan
    mask = ~np.eye(n, dtype=bool)
    vals = mat[mask]
    vals = vals[np.isfinite(vals)]
    if vals.size == 0:
        return np.nan
    if _is_pvalue_method(variant):
        # p-value: меньше лучше -> переводим в "evidence" (больше лучше)
        vals = np.clip(vals, 1e-12, 1.0)
        return float(np.mean(-np.log10(vals)))
    return float(np.mean(np.abs(vals)))


def get_method_spec(variant: str) -> MethodSpec:
    """Совместимость со старым API: возвращает MethodSpec на основе новых семантических множеств."""
    return MethodSpec(
        directed=is_directed_method(variant),
        is_p_value=is_pvalue_method(variant),
        control_dependent=is_control_sensitive_method(variant),
        supports_lag=is_directed_method(variant),
    )


def _residualize_series(y: np.ndarray, X: np.ndarray) -> np.ndarray:
    """Возвращает остатки регрессии y ~ X (с константой)."""
    y = np.asarray(y, dtype=float)
    if X is None or np.size(X) == 0:
        return y - np.nanmean(y)
    X = np.asarray(X, dtype=float)

    # синхронно выкидываем nan
    mask = np.isfinite(y)
    if X.ndim == 1:
        mask &= np.isfinite(X)
    else:
        mask &= np.all(np.isfinite(X), axis=1)

    y2 = y[mask]
    X2 = X[mask]
    if y2.size < 5:
        return y - np.nanmean(y)

    if X2.ndim == 1:
        X2 = X2.reshape(-1, 1)
    X2 = np.column_stack([np.ones(len(X2)), X2])

    try:
        beta, *_ = np.linalg.lstsq(X2, y2, rcond=None)
        resid = np.full_like(y, np.nan, dtype=float)
        resid[mask] = y2 - (X2 @ beta)
        m = np.nanmean(resid)
        resid = np.where(np.isfinite(resid), resid, m)
        return resid
    except Exception:
        return y - np.nanmean(y)


def _pairwise_partial_value(
    data: pd.DataFrame,
    i: int,
    j: int,
    controls: List[int],
    *,
    metric: str,
    lag: int = 1,
) -> float:
    """
    Частичная мера между (i,j) с индивидуальным набором контролей (по индексам колонок).
    metric: 'corr' | 'h2' | 'mi'
    Используем directed-сдвиг: x[t] -> y[t+lag].
    """
    x = data.iloc[:, i].to_numpy(dtype=float)
    y = data.iloc[:, j].to_numpy(dtype=float)

    lag = max(1, int(lag or 1))
    if lag >= len(x) or lag >= len(y):
        return np.nan

    x2 = x[:-lag]
    y2 = y[lag:]

    if controls:
        Z = data.iloc[:, controls].to_numpy(dtype=float)
        Z2 = Z[:-lag, :]
        rx = _residualize_series(x2, Z2)
        ry = _residualize_series(y2, Z2)
    else:
        rx = x2 - np.nanmean(x2)
        ry = y2 - np.nanmean(y2)

    if rx.size < 2 or ry.size < 2:
        return np.nan

    if metric == "corr":
        return float(np.corrcoef(rx, ry)[0, 1])
    if metric == "h2":
        c = float(np.corrcoef(rx, ry)[0, 1])
        return float(c * c)
    if metric == "mi":
        a = rx.reshape(-1, 1)
        b = ry.reshape(-1, 1)
        return float(_knn_mutual_info(a, b, k=DEFAULT_K_MI))
    return np.nan


def _pairwise_partial_matrix(
    data: pd.DataFrame,
    *,
    metric: str,
    lag: int,
    policy: str,
    custom_controls: Optional[List[str]] = None,
) -> np.ndarray:
    """
    NxN матрица частичной меры, где control-set выбирается ПО ПАРЕ.
    policy:
      - 'others': контролируем все остальные переменные (кроме пары)
      - 'custom': контролируем выбранный список custom_controls (пересечённый с остальными)
      - 'none'  : без контролей
    """
    cols = list(data.columns)
    n = len(cols)
    M = np.full((n, n), np.nan, dtype=float)
    for i in range(n):
        M[i, i] = 1.0

    name_to_idx = {c: k for k, c in enumerate(cols)}
    custom_controls = list(custom_controls) if custom_controls else []
    custom_idx = [name_to_idx[c] for c in custom_controls if c in name_to_idx]

    for i in range(n):
        for j in range(i + 1, n):
            if policy == "others":
                controls = [k for k in range(n) if k not in (i, j)]
            elif policy == "custom":
                controls = [k for k in custom_idx if k not in (i, j)]
            elif policy == "none":
                controls = []
            else:
                controls = []
            v = _pairwise_partial_value(data, i, j, controls, metric=metric, lag=lag)
            M[i, j] = v
            M[j, i] = v
    return M

def compute_connectivity_variant(
    data,
    variant,
    lag=1,
    control=None,
    *,
    pairs: Optional[list[tuple[int, int]]] = None,
    partial_mode: str = "global",
    pairwise_policy: str = "others",
    custom_controls: Optional[List[str]] = None,
    control_strategy: str = "none",
    control_pca_k: int = 0,
):
    """
    variant: ключ из method_mapping.
    control: список колонок для GLOBAL partial (как раньше).
    partial_mode:
      - "global"   : прежнее поведение (control применяется ко всей матрице)
      - "pairwise" : control-set выбирается по паре (для *_partial методов, где это реализовано)
    pairwise_policy (для partial_mode="pairwise"):
      - "others" : контролируем все остальные переменные (по умолчанию)
      - "custom" : контролируем custom_controls (пересечение с остальными)
      - "none"   : без контролей
    """
    try:
        if control is not None and len(control) == 0:
            control = None

        # Универсальные «контроли» для partial: global mean / trend / PCA.
        # Это важно для больших N (например, 165 вокселей), где "контроль = все остальные"
        # превращается в плохо интерпретируемый и вычислительно тяжёлый режим.
        control_matrix = None
        control_desc: list[str] = []
        if isinstance(variant, str) and variant.endswith("_partial") and control is None and control_strategy != "none":
            try:
                n = int(len(data))
                ctrls = []
                if control_strategy in {"global", "global_mean", "global_mean_trend", "mean_trend"}:
                    gm = pd.to_numeric(data.mean(axis=1), errors="coerce").to_numpy(dtype=np.float64)
                    ctrls.append(gm)
                    control_desc.append("global_mean")
                if control_strategy in {"global_mean_trend", "mean_trend", "trend"}:
                    t = np.arange(n, dtype=np.float64)
                    t = (t - t.mean()) / (t.std() + 1e-12) if n > 1 else t
                    ctrls.append(t)
                    control_desc.append("linear_trend")
                k = int(max(0, control_pca_k))
                if k > 0:
                    X = data.to_numpy(dtype=np.float64)
                    X = np.nan_to_num(X, nan=0.0)
                    # SVD: time×features
                    U, S, _Vt = np.linalg.svd(X, full_matrices=False)
                    kk = int(min(k, U.shape[1]))
                    if kk > 0:
                        for i in range(kk):
                            ctrls.append(U[:, i] * S[i])
                            control_desc.append(f"pca[{i+1}]")
                if ctrls:
                    control_matrix = np.vstack(ctrls).T
            except Exception:
                control_matrix = None
                control_desc = []

        # pairwise-partial (нужно для N=3..4)
        if partial_mode == "pairwise" and isinstance(variant, str) and variant.endswith("_partial"):
            if variant == "correlation_partial":
                return _pairwise_partial_matrix(
                    data,
                    metric="corr",
                    lag=max(1, int(lag or 1)),
                    policy=pairwise_policy,
                    custom_controls=custom_controls,
                )
            if variant == "h2_partial":
                return _pairwise_partial_matrix(
                    data,
                    metric="h2",
                    lag=max(1, int(lag or 1)),
                    policy=pairwise_policy,
                    custom_controls=custom_controls,
                )
            if variant == "mutinf_partial":
                return _pairwise_partial_matrix(
                    data,
                    metric="mi",
                    lag=max(1, int(lag or 1)),
                    policy=pairwise_policy,
                    custom_controls=custom_controls,
                )
            # Остальные partial пока считаем GLOBAL-режимом (через control):
            # корректное "pairwise partial" для TE/AH/Granger требует отдельной
            # условной постановки и другого API.

        if variant in method_mapping:
            metric_func = get_metric_func(variant)
            return metric_func(
                data,
                lag=lag,
                control=control,
                control_matrix=control_matrix,
                control_desc=control_desc,
                pairs=pairs,
            )

        return connectivity.correlation_matrix(data)
    except Exception as e:
        logging.error(f"[ComputeVariant] Метод {variant} не работает: {e}")
        return None


def powerset(iterable):
    s = list(iterable)
    return chain.from_iterable(combinations(s, r) for r in range(len(s)+1))

##############################################
#  диагностика коэффициентов регрессии
##############################################
def regression_diagnostics(df: pd.DataFrame, target: str, controls: list):
    """
    Рассчитывает линейную регрессию вида: target ~ controls.
    Возвращает строку с R².
    """
    # Если нет контрольных переменных — выходим
    if not controls:
        return f"Нет контрольных переменных для {target}."
    # Иначе строим модель и возвращаем R²
    X = df[controls]
    y = df[target]
    model = LinearRegression().fit(X, y)
    r2 = model.score(X, y)
    return f"{target} ~ {controls}: R² = {r2:.3f}"


##############################################
# Частотный анализ: возвращает пиковые значения
##############################################
def frequency_analysis(series: pd.Series, peak_height_ratio: float = 0.2, *, fs: float = 1.0):
    freqs, amplitude, phase, peaks = plt_fft_analysis(series, fs=fs)
    if freqs.size == 0 or peaks.size == 0:
        return None, None, None
    peak_freqs = freqs[peaks]
    peak_amps = amplitude[peaks]
    periods = 1 / peak_freqs
    return peak_freqs, peak_amps, periods

def sliding_fft_analysis(data: pd.DataFrame, window_size: int, overlap: int) -> dict:
    """Экспериментальный анализ скользящего FFT (по умолчанию отключён)."""
    logging.info("[Sliding FFT] Экспериментальная функция отключена.")
    return {}


def analyze_sliding_windows_with_metric(
    data: pd.DataFrame,
    variant: str,
    window_size: int,
    stride: int,
    *,
    lag: int = 1,
    pairs: Optional[list[tuple[int, int]]] = None,
    start_min: int | None = None,
    start_max: int | None = None,
    max_windows: int = 400,
    return_matrices: bool = False,
) -> dict:
    """Анализ скользящих окон для заданного window_size.

    Возвращает структуру для HTML-отчёта:
      {
        "best_window": {"start": int, "end": int, "metric": float, "matrix": np.ndarray},
        "curve": {"x": [start_idx...], "y": [metric...]},
        "ticks": [{"start":..,"end":..,"metric":..,"matrix":..|None}, ...],
        "extremes": {"best": idx|None, "median": idx|None, "worst": idx|None}
      }

    stride — шаг сдвига окна (в точках), lag — лаг (только для lagged-методов).
    """
    if data is None or data.empty:
        return {}

    n = len(data)
    w = int(max(2, min(window_size, n)))
    s = int(max(1, stride))

    xs: List[int] = []
    ys: List[float] = []
    ticks: list[dict] = []

    best = {
        "start": 0,
        "end": w,
        "metric": float("-inf"),
        "matrix": None,
    }

    # Диапазон позиций окна (включительно по start, end=start+w).
    st0 = int(max(0, start_min)) if start_min is not None else 0
    st1 = int(min(n - w, start_max)) if start_max is not None else (n - w)
    st1 = int(max(st0, st1))

    # Ограничиваем количество окон, чтобы не взорваться по времени на длинных рядах.
    # Это не запрещает пользователю уменьшить stride, но даёт безопасный дефолт.
    max_windows = int(max(1, max_windows))
    starts = list(range(st0, st1 + 1, s))
    if len(starts) > max_windows:
        # равномерная подвыборка
        idx = np.linspace(0, len(starts) - 1, max_windows).round().astype(int)
        starts = [starts[i] for i in idx]

    for start in starts:
        end = start + w
        if end > n:
            break
        chunk = data.iloc[start:end]
        try:
            mat = compute_connectivity_variant(chunk, variant, lag=int(max(1, lag)), pairs=pairs)
            score = _lag_quality(variant, mat)
        except Exception as ex:
            logging.error(f"[SlidingWindow] {variant} win={w} start={start}: {ex}")
            mat = None
            score = float("nan")

        xs.append(int(start))
        ys.append(float(score) if np.isfinite(score) else float("nan"))

        ticks.append(
            {
                "start": int(start),
                "end": int(end),
                "metric": float(score) if np.isfinite(score) else float("nan"),
                "matrix": mat if return_matrices else None,
            }
        )

        if np.isfinite(score) and float(score) > float(best["metric"]):
            best = {
                "start": int(start),
                "end": int(end),
                "metric": float(score),
                "matrix": mat,
            }

    return {
        "best_window": best,
        "curve": {"x": xs, "y": ys},
        "ticks": ticks,
        "extremes": _select_best_median_worst(ticks, key="metric"),
    }


def _select_best_median_worst(items: list[dict], *, key: str = "metric") -> dict:
    """Возвращает индексы best/median/worst по значению key (с учётом NaN)."""
    if not items:
        return {"best": None, "median": None, "worst": None}

    vals: list[tuple[int, float]] = []
    for i, it in enumerate(items):
        try:
            v = float(it.get(key, float("nan")))
        except Exception:
            v = float("nan")
        if np.isfinite(v):
            vals.append((i, v))

    if not vals:
        return {"best": None, "median": None, "worst": None}

    vals_sorted = sorted(vals, key=lambda t: t[1])
    worst_i = int(vals_sorted[0][0])
    best_i = int(vals_sorted[-1][0])
    med_val = float(np.median([v for _, v in vals_sorted]))
    median_i = int(min(vals_sorted, key=lambda t: abs(t[1] - med_val))[0])
    return {"best": best_i, "median": median_i, "worst": worst_i}


def sliding_window_pairwise_analysis(
    data: pd.DataFrame,
    method: str,
    window_size: int,
    overlap: int,
) -> dict:
    """Экспериментальный парный анализ скользящих окон (по умолчанию отключён)."""
    logging.info("[Sliding Pairwise] Экспериментальная функция отключена.")
    return {}

##############################################
# Листы с коэффициентами и частотным анализом
##############################################
def export_coefficients_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Coefficients & Explanations")
    ws.append(["Описание:", "Лист содержит краткие пояснения коэффициентов регрессий и матриц связей."])
    ws.append(["Например, коэффициенты регрессии показывают, как контрольные переменные влияют на связь между переменными."])
    ws.append([])
    ws.append(["Регрессионная диагностика:"])
    ws.append(["Переменная", "Контроль", "Диагностика"])
    for target in tool.data.columns:
        controls = [c for c in tool.data.columns if c != target]
        diag_str = regression_diagnostics(tool.data, target, controls)
        ws.append([target, str(controls), diag_str])
    ws.append([])
    ws.append(["Матрицы связей:"])
    ws.append(["Метод", "Описание"])
    methods_info = [
        ("correlation_full", "Стандартная корреляционная матрица."),
        ("correlation_partial", "Частичная корреляция (с контролем)."),
        ("mutinf_full", "Полная взаимная информация."),
        ("coherence_full", "Когерентность между переменными.")
    ]
    for m, info in methods_info:
        ws.append([m, info])
    logging.info("[Coefficients] Лист 'Coefficients & Explanations' сформирован.")

def export_frequency_summary_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Frequency Summary")
    ws.append(["Столбец", "Пиковые частоты", "Пиковые амплитуды", "Периоды", "Пояснение"])
    for col in tool.data.columns:
        s = tool.data[col].dropna()
        freq, amps, periods = frequency_analysis(s)
        if freq is not None:
            freq_str = ", ".join([f"{f:.3f}" for f in freq])
            amps_str = ", ".join([f"{f:.3f}" for f in amps])
            period_str = ", ".join([f"{p:.1f}" for p in periods])
            note = f"Макс. связь на {freq[np.argmax(amps)]:.3f} Hz"
        else:
            freq_str = amps_str = period_str = "Нет пиков"
            note = "Пиковые частоты не выявлены"
        ws.append([col, freq_str, amps_str, period_str, note])
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length
    logging.info("[Frequency] Лист 'Frequency Summary' сформирован.")

##############################################
# Новый лист: Индивидуальные АЧХ и ФЧХ (раздельно)
##############################################
def export_individual_ac_ph_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Individual AC & PH")
    ws.append(["Столбец", "АЧХ", "ФЧХ"])
    plots = plot_individual_ac_ph(tool.data_normalized, "Individual AC & PH")
    for col, imgs in plots.items():
        ws.append([col])
        img_ac = Image(imgs["AC"])
        img_ac.width = 400
        img_ac.height = 300
        ws.add_image(img_ac, f"B{ws.max_row}")
        img_ph = Image(imgs["PH"])
        img_ph.width = 400
        img_ph.height = 300
        ws.add_image(img_ph, f"G{ws.max_row}")
    logging.info("[Individual AC & PH] Лист сформирован.")

##############################################
# Новый лист: Анализ энтропии
##############################################
def export_entropy_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Entropy Analysis")
    ws.append(["Столбец", "Sample Entropy (sampen)"])
    for col in tool.data.columns:
        s = tool.data[col].dropna()
        ent = compute_sample_entropy(s)
        ws.append([col, f"{ent:.3f}" if not np.isnan(ent) else "N/A"])
    logging.info("[Entropy Analysis] Лист сформирован.")

##############################################
# Новый лист
##############################################
def export_combined_informational_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Combined Informational Analysis")
    current_row = 1
    ws.cell(row=current_row, column=1, value="Combined Informational Analysis")
    current_row += 2
    ws.cell(row=current_row, column=1, value="Lag Analysis Summary (Aggregated)")
    current_row += 1
    buf_lag = tool.plot_all_methods_lag_comparison(tool.lag_results)
    img_lag = Image(buf_lag)
    img_lag.width = 800
    img_lag.height = 600
    ws.add_image(img_lag, f"A{current_row}")
    current_row += 30
    ws.cell(row=current_row, column=1, value="Sliding Window Analysis Summary (Aggregated)")
    current_row += 1
    sw_res = tool.analyze_sliding_windows(
        "coherence_full",
        window_size=min(50, len(tool.data_normalized) // 2),
        overlap=min(25, len(tool.data_normalized) // 4),
    )
    if sw_res:
        legend_text = "Метод: coherence_full, Окно: 50"
        buf_sw = tool.plot_sliding_window_comparison(sw_res, legend_text=legend_text)
        img_sw = Image(buf_sw)
        img_sw.width = 700
        img_sw.height = 400
        ws.add_image(img_sw, f"A{current_row}")
        current_row += 20
    else:
        ws.append(["Sliding Window Analysis отключён или нет данных."])
        current_row += 2
    ws.cell(row=current_row, column=1, value="Pairwise Lag Analysis (пример для первой пары)")
    current_row += 1
    if len(tool.data.columns) >= 2:
        pair = list(combinations(tool.data.columns, 2))[0]
        col1, col2 = pair
        series1 = tool.data[col1].dropna().values
        series2 = tool.data[col2].dropna().values
        n = min(len(series1), len(series2))
        lag_metrics = {}
        for lag in range(1, 21):
            if n > lag:
                corr = np.corrcoef(series1[lag:], series2[:n-lag])[0, 1] if len(series1[lag:]) > 1 and len(series2[:n-lag]) > 1 else np.nan
                lag_metrics[lag] = corr
        if lag_metrics:
            lags = list(lag_metrics.keys())
            correlations = [lag_metrics[lag] for lag in lags]
            fig, ax = plt.subplots(figsize=(4, 3))
            ax.plot(lags, correlations, marker='o')
            ax.set_title(f"Lag Analysis: {col1}-{col2}")
            legend_text_pair = f"Пара: {col1}-{col2}, Метод: Lag Correlation"
            ax.text(0.5, 0.1, legend_text_pair, transform=ax.transAxes, fontsize=8, 
                    verticalalignment='bottom', bbox=dict(facecolor='white', alpha=0.5))
            buf_plag = BytesIO()
            plt.tight_layout()
            plt.savefig(buf_plag, format="png", dpi=100)
            buf_plag.seek(0)
            plt.close(fig)
            img_plag = Image(buf_plag)
            img_plag.width = 400
            img_plag.height = 300
            ws.add_image(img_plag, f"A{current_row}")
        else:
            ws.append(["Недостаточно данных для Pairwise Lag Analysis."])
        current_row += 20
        # (пример для первой пары)
        ws.cell(row=current_row, column=1, value="Extended Spectral Analysis (пример для первой пары)")
        current_row += 1
        title_es = f"Coherence {col1}-{col2}"
        buf_es = plot_coherence_vs_frequency(tool.data[col1], tool.data[col2], title_es, fs=getattr(tool, "fs", 1.0))
        img_es = Image(buf_es)
        img_es.width = 400
        img_es.height = 300
        ws.add_image(img_es, f"A{current_row}")
        current_row += 20
        # (пример для первой пары)
        ws.cell(row=current_row, column=1, value="Frequency Demonstration (пример для первой пары)")
        current_row += 1
        buf_fd = plot_coherence_vs_frequency(tool.data[col1], tool.data[col2], title_es, fs=getattr(tool, "fs", 1.0))
        img_fd = Image(buf_fd)
        img_fd.width = 400
        img_fd.height = 300
        ws.add_image(img_fd, f"A{current_row}")
        current_row += 20
    else:
        ws.append(["Недостаточно столбцов для Pairwise Lag Analysis или Spectral Analysis."])
        current_row += 60 
    ws.cell(row=current_row, column=1, value="End of Combined Informational Analysis")
    logging.info("[Combined Informational Analysis] Лист сформирован.")

##############################################
#Combined Time Series (агрегированный + индивидуальные графики)
##############################################
def export_combined_ts_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Combined Time Series")
    ws.append(["Aggregated Time Series: Оригинальные и Нормализованные (на одном графике)"])
    buf_orig = tool.plot_time_series(tool.data, "Aggregated Original Time Series")
    img_orig = Image(buf_orig)
    img_orig.width = 600
    img_orig.height = 300
    ws.add_image(img_orig, "A2")
    
    ws.append([])
    buf_norm = tool.plot_time_series(tool.data_normalized, "Aggregated Normalized Time Series")
    img_norm = Image(buf_norm)
    img_norm.width = 600
    img_norm.height = 300
    ws.add_image(img_norm, "A10")
    
    ws.append(["Individual Time Series Plots:"])
    row = ws.max_row + 2
    for col in tool.data.columns:
        buf_ind = tool.plot_single_time_series(tool.data[col], f"Original: {col}")
        img_ind = Image(buf_ind)
        img_ind.width = 300
        img_ind.height = 200
        ws.add_image(img_ind, f"A{row}")
        buf_ind_norm = tool.plot_single_time_series(tool.data_normalized[col], f"Normalized: {col}")
        img_ind_norm = Image(buf_ind_norm)
        img_ind_norm.width = 300
        img_ind_norm.height = 200
        ws.add_image(img_ind_norm, f"E{row}")
        row += 15
    logging.info("[Combined Time Series] Лист сформирован.")

##############################################
# Combined FFT (агрегированный + индивидуальные графики)
##############################################
def export_all_fft_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Combined FFT")
    ws.append(["Combined FFT Analysis (Aggregated) - Original"])
    buf_fft_orig = tool.plot_fft(tool.data, "Aggregated Original FFT")
    img_fft_orig = Image(buf_fft_orig)
    img_fft_orig.width = 600
    img_fft_orig.height = 400
    ws.add_image(img_fft_orig, "A2")
    
    ws.append([])
    ws.append(["Combined FFT Analysis (Aggregated) - Normalized"])
    buf_fft_norm = tool.plot_fft(tool.data_normalized, "Aggregated Normalized FFT")
    img_fft_norm = Image(buf_fft_norm)
    img_fft_norm.width = 600
    img_fft_norm.height = 400
    ws.add_image(img_fft_norm, "A20")
    
    ws.append(["Individual FFT Analysis:"])
    row = ws.max_row + 2
    for col in tool.data.columns:
        buf_fft_ind = tool.plot_single_fft(tool.data[col], f"Original FFT: {col}")
        img_fft_ind = Image(buf_fft_ind)
        img_fft_ind.width = 300
        img_fft_ind.height = 200
        ws.add_image(img_fft_ind, f"A{row}")
        buf_fft_ind_norm = tool.plot_single_fft(tool.data_normalized[col], f"Normalized FFT: {col}")
        img_fft_ind_norm = Image(buf_fft_ind_norm)
        img_fft_ind_norm.width = 300
        img_fft_ind_norm.height = 200
        ws.add_image(img_fft_ind_norm, f"E{row}")
        row += 15
    logging.info("[Combined FFT] Лист сформирован.")

##############################################
# Создание оглавления с гиперссылками
##############################################
def create_table_of_contents(wb: Workbook):
    if "Table of Contents" in wb.sheetnames:
        old_sheet = wb["Table of Contents"]
        wb.remove(old_sheet)
    toc = wb.create_sheet("Table of Contents", 0)
    row = 1
    for sheet_name in wb.sheetnames:
        if sheet_name == "Table of Contents":
            continue
        link = f"#{sheet_name}!A1"
        cell = toc.cell(row=row, column=1)
        cell.value = sheet_name
        cell.hyperlink = link
        cell.style = "Hyperlink"
        row += 1

##############################################
# Класс BigMasterTool 
##############################################
class BigMasterTool:
    """
    Главный класс-оркестратор.
    Управляет загрузкой данных, запуском расчетов и вызовом генераторов отчетов.
    """

    def __init__(
        self,
        data: pd.DataFrame = None,
        enable_experimental: bool = False,
        config: Optional[AnalysisConfig] = None,
        stage_callback=None,
    ) -> None:
        # Внутренний лог пайплайна (для UI/отчётов). Не путать с модулем logging.
        self.log = RunLog()

        # Инициализация данных
        # Снимки для отчёта: raw -> preprocessed -> after auto-diff.
        self.data_raw = pd.DataFrame()
        self.data_preprocessed = pd.DataFrame()
        self.data_after_autodiff = pd.DataFrame()
        self.preprocessing_report = None  # type: ignore
        self.autodiff_report = {"enabled": False, "differenced": []}
        # Уменьшение размерности: снимок состояния текущего запуска.
        self.dimred_report = {"enabled": False, "method": "none"}
        self.dimred_mapping: pd.DataFrame = pd.DataFrame()
        self.data_dimred: pd.DataFrame = pd.DataFrame()
        self.data_dimred_base: Optional[pd.DataFrame] = None
        self.pairwise_summaries = {}
        if data is not None:
            # Чистка константных колонок
            data = data.loc[:, (data != data.iloc[0]).any()]
            self.data = data.copy()
            # ВАЖНО: матрица рядов не должна таскать тяжелые attrs (coords и пр.).
            # Иначе при self.data[col] pandas может сделать deepcopy(attrs) и резко увеличить память.
            try:
                self.data.attrs = {}
            except Exception:
                try:
                    self.data.attrs.clear()
                except Exception:
                    pass
            # Приведение к числам
            for c in list(self.data.columns):
                self.data[c] = pd.to_numeric(self.data[c], errors="coerce")
            # Если нет имен колонок, даем стандартные c1, c2...
            if len(self.data.columns) > 0 and isinstance(self.data.columns[0], int):
                self.data.columns = [f"c{i + 1}" for i in range(self.data.shape[1])]
        else:
            self.data = pd.DataFrame()

        self.data_normalized: pd.DataFrame = pd.DataFrame()
        self.results: dict = {}
        self.results_meta: dict = {}
        self.variant_lags: dict = {}
        self.window_analysis: dict = {}
        self.config: AnalysisConfig = config or AnalysisConfig(enable_experimental=enable_experimental)

        # Коллбек прогресса/этапов для UI/CLI.
        # Сигнатура: cb(stage: str, progress: float|None, meta: dict)
        self._stage_callback = stage_callback

        # Настройки
        self.fs: float = 1.0  # Частота дискретизации

        # Определяем списки методов
        self.undirected_methods = [m for m in method_mapping if not get_method_spec(m).directed]
        self.directed_methods = [m for m in method_mapping if get_method_spec(m).directed]

    def set_stage_callback(self, cb) -> None:
        """Установить коллбек прогресса/этапов (для GUI/Web/CLI)."""
        self._stage_callback = cb

    def _stage(self, stage: str, progress: Optional[float] = None, **meta) -> None:
        """Сообщить текущий этап. progress ∈ [0..1] или None."""
        try:
            logging.info("[Stage] %s", stage)
        except Exception:
            pass
        cb = getattr(self, "_stage_callback", None)
        if cb is None:
            return
        try:
            cb(stage, progress, dict(meta or {}))
        except Exception:
            # UI не должен валить вычисления.
            return

    def load_data_excel(self, filepath: str, **kwargs) -> pd.DataFrame:
        """Загружает данные, чистит их и опционально устраняет нестационарность."""
        self._stage("Загрузка данных", 0.0, file=str(filepath))
        qc_enabled = bool(kwargs.pop("qc_enabled", True))
        # 1) Сырой numeric-слепок без предобработки для честного before/after в отчёте.
        try:
            self._stage("Загрузка RAW (без предобработки)", 0.05)
            self.data_raw = load_or_generate(
                filepath,
                preprocess=False,
                normalize=False,
                remove_outliers=False,
                fill_missing=False,
                check_stationarity=False,
            )
        except Exception:
            self.data_raw = pd.DataFrame()

        # 2) Основной путь загрузки (с учётом выбранных опций) + структурированный отчёт.
        self._stage("Загрузка + предобработка", 0.15)
        df_out = load_or_generate(filepath, return_report=True, **kwargs)
        if isinstance(df_out, tuple):
            self.data, self.preprocessing_report = df_out
        else:
            self.data, self.preprocessing_report = df_out, None
        self.data_preprocessed = self.data.copy()

        self._stage("Данные загружены", 0.35, shape=list(self.data.shape))

        # Координаты (для voxel-wide формата) и QC
        self.coords_df = None
        self.qc_raw = None
        self.qc_clean = None
        try:
            notes = (self.preprocessing_report.notes if self.preprocessing_report is not None else {}) or {}
            coords = notes.get("coords")
            if isinstance(coords, list) and coords:
                self.coords_df = pd.DataFrame(coords)
        except Exception:
            self.coords_df = None

        # QC только если включено и (есть много рядов или есть coords)
        try:
            from ..analysis import stats as _stats
            if qc_enabled and (self.coords_df is not None or int(self.data.shape[1]) >= 20):
                self._stage("QC (качество вокселей/рядов)", 0.45)
                if getattr(self, "data_raw", pd.DataFrame()).empty is False:
                    self.qc_raw = _stats.voxel_qc(self.data_raw, coords=self.coords_df)
                self.qc_clean = _stats.voxel_qc(self.data, coords=self.coords_df)
        except Exception:
            self.qc_raw = None
            self.qc_clean = None

        # Запоминаем настройку QC (нужно для отчётов/пояснений)
        try:
            self.results_meta.setdefault("__run__", {})
            self.results_meta["__run__"].setdefault("qc_enabled", qc_enabled)
        except Exception:
            pass

        self.data = self.data.fillna(self.data.mean(numeric_only=True))

        if self.config.auto_difference:
            self.autodiff_report = {"enabled": True, "differenced": []}
            self._stage("Проверка стационарности + авто-дифференцирование", 0.55)
            logging.info("Запущена проверка стационарности и авто-дифференцирование...")
            diff_count = 0
            for col in self.data.columns:
                if not pd.api.types.is_numeric_dtype(self.data[col]):
                    continue

                stat, pval = analysis_stats.test_stationarity(self.data[col])
                if stat is None:
                    continue
                if pval is not None and pval > 0.05:
                    self.data[col] = self.data[col].diff().fillna(0)
                    diff_count += 1
                    self.autodiff_report.setdefault("differenced", []).append(col)

            if diff_count > 0:
                logging.warning("Применено дифференцирование к %s нестационарным рядам.", diff_count)

        self.data_after_autodiff = self.data.copy()
        logging.info("[BigMasterTool] Данные готовы: %s", self.data.shape)
        self._stage("Данные готовы", 0.70, shape=list(self.data.shape))
        return self.data

    def normalize_data(self) -> None:
        """Нормализация данных (Z-score)."""
        if self.data.empty:
            logging.warning("Нет данных для нормализации.")
            return

        self._stage("Нормализация", 0.75)

        cols = [c for c in self.data.columns if pd.api.types.is_numeric_dtype(self.data[c])]
        scaler = StandardScaler()
        self.data_normalized = self.data.copy()
        self.data_normalized[cols] = scaler.fit_transform(self.data[cols])
        logging.info("Данные нормализованы.")

    def _apply_fdr_correction(self) -> None:
        """Применяет поправку Бенджамини-Хохберга к p-value методам."""
        if self.config.pvalue_correction != "fdr_bh":
            return

        for variant, mat in self.results.items():
            if mat is None or not is_pvalue_method(variant):
                continue
            self.results[variant] = apply_pvalue_correction_matrix(
                mat,
                directed=get_method_spec(variant).directed,
            )
            logging.info("Применена FDR коррекция к %s", variant)

    def run_all_methods(self, **kwargs) -> None:
        """Запуск всех доступных методов анализа."""
        self._stage("Подготовка: dimred/нормализация", 0.72)
        # Опционально уменьшаем размерность до нормализации/оценки связности.
        self._maybe_apply_dimred(**kwargs)
        self._maybe_post_preprocess(**kwargs)
        self.normalize_data()
        if self.data_normalized.empty:
            return

        prev_run_meta = dict((getattr(self, "results_meta", {}) or {}).get("__run__", {}) or {})
        self.results = {}
        # Метаданные (лаг, окна, partial-контроль и т.п.) — чтобы отчёты могли
        # объяснять пользователю «что именно было исключено/подобрано».
        self.results_meta = {}
        if prev_run_meta:
            self.results_meta["__run__"] = prev_run_meta
        self.variant_lags = {}
        self.window_analysis = {}

        variants_all = list(method_mapping.keys())
        n_total = max(1, len(variants_all))
        for idx, variant in enumerate(variants_all, start=1):
            self._stage(
                f"Расчёт: {variant} ({idx}/{n_total})",
                0.80 + 0.18 * (idx - 1) / n_total,
                variant=variant,
                i=idx,
                n=n_total,
            )
            logging.info("Расчет метода: %s...", variant)
            try:
                mat, meta = self._compute_variant_auto(variant, **kwargs)
                self.results[variant] = mat
                self.results_meta[variant] = meta
                if meta.get("chosen_lag") is not None:
                    self.variant_lags[variant] = int(meta["chosen_lag"])
                if meta.get("window") is not None:
                    self.window_analysis[variant] = meta["window"]
            except Exception as e:
                logging.error("Ошибка в методе %s: %s", variant, e)
                self.results[variant] = None
                self.results_meta[variant] = {"error": str(e)}

        self._apply_fdr_correction()
        self._stage("Готово: расчёты завершены", 1.0)
        logging.info("Все расчеты завершены.")

    def run_selected_methods(self, variants: List[str], max_lag: int = 5, **kwargs) -> Dict[str, int]:
        """
        Запуск конкретных выбранных методов.
        Возвращает словарь {метод: использованный_лаг}.
        """
        self._stage("Подготовка: dimred/нормализация", 0.72)
        # Опционально уменьшаем размерность до нормализации/оценки связности.
        self._maybe_apply_dimred(**kwargs)
        self._maybe_post_preprocess(**kwargs)
        self.normalize_data()
        prev_run_meta = dict((getattr(self, "results_meta", {}) or {}).get("__run__", {}) or {})
        self.results = {}
        self.results_meta = {}
        if prev_run_meta:
            self.results_meta["__run__"] = prev_run_meta
        self.window_analysis = {}

        # сохраняем параметры запуска (для UI/CLI пояснений)
        try:
            self.results_meta.setdefault("__run__", {})
            self.results_meta["__run__"].update(
                {
                    "variants": list(variants),
                    "max_lag": int(max_lag),
                    "lag_selection": (kwargs.get("lag_selection") or self.config.lag_selection),
                    "lag": kwargs.get("lag"),
                    "window_sizes": kwargs.get("window_sizes"),
                    "window_stride": kwargs.get("window_stride"),
                    "window_policy": kwargs.get("window_policy"),
                    "control_strategy": kwargs.get("control_strategy", "none"),
                    "control_pca_k": int(kwargs.get("control_pca_k", 0) or 0),
                    "dimred": dict(getattr(self, "dimred_report", {}) or {}),
                }
            )
        except Exception:
            pass

        # max_lag аргументом оставляем, но если задан self.config.max_lag —
        # берём максимум из них (чтобы не ломать старые вызовы).
        self.config.max_lag = int(max(self.config.max_lag, int(max_lag)))

        method_options = kwargs.get("method_options") or {}
        if method_options is None:
            method_options = {}

        used_lags: Dict[str, int] = {}
        n_total = max(1, len(variants))
        for idx, variant in enumerate(variants, start=1):
            self._stage(
                f"Расчёт: {variant} ({idx}/{n_total})",
                0.80 + 0.18 * (idx - 1) / n_total,
                variant=variant,
                i=idx,
                n=n_total,
            )
            if variant not in method_mapping:
                continue
            try:
                v_kwargs = dict(kwargs)
                if isinstance(method_options, dict) and isinstance(method_options.get(variant), dict):
                    # Локальные опции метода перекрывают глобальные kwargs.
                    v_kwargs.update(method_options.get(variant) or {})
                mat, meta = self._compute_variant_auto(variant, **v_kwargs)
                self.results[variant] = mat
                self.results_meta[variant] = meta
                if meta.get("chosen_lag") is not None:
                    used_lags[variant] = int(meta["chosen_lag"])
                if meta.get("window") is not None:
                    self.window_analysis[variant] = meta["window"]
            except Exception as e:
                logging.error("Ошибка %s: %s", variant, e)
                self.results[variant] = None
                self.results_meta[variant] = {"error": str(e)}

        self.variant_lags = used_lags
        self._stage("Готово: расчёты завершены", 1.0)
        return used_lags

    def calculate_graph_metrics(self, threshold: float = 0.2) -> None:
        """[PATCH] Рассчитать топологические метрики для уже посчитанных матриц.

        Метод безопасен для повторного вызова: каждый запуск перезаписывает
        ``self.graph_results`` и анализирует только непустые матрицы из ``self.results``.
        """
        from ..analysis.graph import analyze_graph_topology

        self.graph_results = {}
        names = list(self.data.columns)

        for variant, mat in (self.results or {}).items():
            if mat is None:
                continue

            is_directed = get_method_spec(variant).directed

            # Для p-value матриц: строим "силу связи" как (1 - p),
            # но только для статистически значимых пар p < threshold.
            if is_pvalue_method(variant):
                clean_mat = np.zeros_like(mat, dtype=float)
                mask = (mat < float(threshold)) & (~np.eye(mat.shape[0], dtype=bool))
                clean_mat[mask] = 1.0 - mat[mask]
                analysis = analyze_graph_topology(clean_mat, names, threshold=0.01, directed=is_directed)
            else:
                analysis = analyze_graph_topology(mat, names, threshold=float(threshold), directed=is_directed)

            self.graph_results[variant] = analysis
            logging.info("[Graph] Analyzed topology for %s", variant)

    def _compute_variant_auto(self, variant: str, **kwargs) -> tuple[np.ndarray | None, dict]:
        """Единая точка расчёта: лаг (fixed/optimize) + окна (none/best/mean).

        Архитектуру сохраняем: наружу по-прежнему отдаём матрицу, но параллельно
        собираем meta для отчётов.
        """
        df = self.data_normalized
        if df is None or df.empty:
            return None, {"error": "empty data"}

        # --- partial controls description ---
        meta: dict = {"variant": variant}
        if is_control_sensitive_method(variant):
            # В текущей реализации (pairwise_policy='others' по умолчанию) это значит:
            # для пары (Xi, Xj) исключаем влияние всех остальных переменных.
            meta["partial"] = {
                "mode": kwargs.get("partial_mode", "pairwise"),
                "pairwise_policy": kwargs.get("pairwise_policy", "others"),
                "custom_controls": kwargs.get("custom_controls"),
                "control_strategy": kwargs.get("control_strategy", "none"),
                "control_pca_k": int(kwargs.get("control_pca_k", 0) or 0),
                "explain": "Для каждой пары (Xi, Xj) исключено линейное влияние набора control.",
            }

        # --- large-N simplification: full matrix vs pairs vs spatial neighbors ---
        n_cols = int(df.shape[1])
        pair_mode = str(kwargs.get("pair_mode") or "auto").lower()
        auto_thr = int(kwargs.get("pair_auto_threshold") or 0)
        if auto_thr <= 0:
            auto_thr = 600

        def _parse_pairs_text(text: str) -> list[tuple[int, int]]:
            """Parse user pairs like: a-b; a->b; 0-1; 0->1.

            Names are matched against df.columns.
            """
            text = (text or "").strip()
            if not text:
                return []
            # normalize separators
            raw_items: list[str] = []
            for token in text.replace("—", "-").replace(",", ";").split(";"):
                tt = token.strip()
                if tt:
                    raw_items.append(tt)
            col_to_idx = {str(c): i for i, c in enumerate(df.columns)}
            pairs_out: list[tuple[int, int]] = []
            for it in raw_items:
                if "->" in it:
                    a, b = [x.strip() for x in it.split("->", 1)]
                elif "-" in it:
                    a, b = [x.strip() for x in it.split("-", 1)]
                else:
                    continue
                def _to_idx(x: str) -> Optional[int]:
                    if x.isdigit():
                        ix = int(x)
                        return ix if 0 <= ix < n_cols else None
                    return col_to_idx.get(x)
                ia = _to_idx(a)
                ib = _to_idx(b)
                if ia is None or ib is None or ia == ib:
                    continue
                pairs_out.append((int(ia), int(ib)))
            return pairs_out

        def _build_neighbor_pairs(kind: str, radius: int) -> list[tuple[int, int]]:
            """Spatial neighborhood pairs from coords_df (voxel_id/x/y/z).

            If coords are missing, returns empty list.
            """
            if getattr(self, "coords_df", None) is None or self.coords_df is None:
                return []
            coords = self.coords_df
            if coords.empty:
                return []
            # map voxel_id -> column index
            col_to_idx = {str(c): i for i, c in enumerate(df.columns)}
            # build coord -> index
            coord_to_idx: dict[tuple[int, int, int], int] = {}
            for _, r in coords.iterrows():
                vid = str(r.get("voxel_id"))
                if vid not in col_to_idx:
                    continue
                try:
                    x = int(r.get("x"))
                    y = int(r.get("y"))
                    z = int(r.get("z"))
                except Exception:
                    continue
                coord_to_idx[(x, y, z)] = int(col_to_idx[vid])

            if not coord_to_idx:
                return []
            kind = str(kind or "26")
            radius = int(max(1, radius))
            pairs_out: list[tuple[int, int]] = []
            # neighbor offsets
            offsets: list[tuple[int, int, int]] = []
            for dx in range(-radius, radius + 1):
                for dy in range(-radius, radius + 1):
                    for dz in range(-radius, radius + 1):
                        if dx == dy == dz == 0:
                            continue
                        if kind == "6":
                            if abs(dx) + abs(dy) + abs(dz) == 1:
                                offsets.append((dx, dy, dz))
                        else:
                            # 26-neighborhood (chebyshev)
                            if max(abs(dx), abs(dy), abs(dz)) <= radius:
                                offsets.append((dx, dy, dz))

            seen = set()
            for (x, y, z), i in coord_to_idx.items():
                for dx, dy, dz in offsets:
                    j = coord_to_idx.get((x + dx, y + dy, z + dz))
                    if j is None or j == i:
                        continue
                    a, b = (i, j) if i < j else (j, i)
                    key = (a, b)
                    if key in seen:
                        continue
                    seen.add(key)
                    pairs_out.append((a, b))
            return pairs_out

        # resolve effective pairs list
        pairs_idx: Optional[list[tuple[int, int]]] = None
        if pair_mode == "auto":
            pair_mode = "neighbors" if n_cols >= auto_thr else "full"
        if pair_mode == "pairs":
            pairs_idx = _parse_pairs_text(str(kwargs.get("pairs_text") or ""))
        elif pair_mode == "neighbors":
            pairs_idx = _build_neighbor_pairs(str(kwargs.get("neighbor_kind") or "26"), int(kwargs.get("neighbor_radius") or 1))
        elif pair_mode == "random":
            # random undirected pairs (fallback when no coords). Deterministic seed.
            max_pairs = int(max(1, kwargs.get("max_pairs") or 50000))
            rng = np.random.default_rng(12345)
            # sample indices in upper triangle
            m = min(max_pairs, max(1, n_cols * 5))
            pairs_out = set()
            while len(pairs_out) < m:
                i = int(rng.integers(0, n_cols))
                j = int(rng.integers(0, n_cols))
                if i == j:
                    continue
                a, b = (i, j) if i < j else (j, i)
                pairs_out.add((a, b))
            pairs_idx = list(pairs_out)
        else:
            pairs_idx = None

        if pairs_idx is not None:
            meta["pair_mode"] = str(pair_mode)
            meta["pairs_count"] = int(len(pairs_idx))
            meta["pairs_explain"] = (
                "Матрица считается упрощённо: только по выбранным парам (остальные пары заполняются дефолтом метода)."
            )

        # --- lag selection ---
        supports_lag = is_directed_method(variant) or variant.startswith("granger") or variant.startswith("te_") or variant.startswith("ah_")
        lag_sel = (kwargs.get("lag_selection") or self.config.lag_selection or "optimize").lower()
        max_lag = int(kwargs.get("max_lag") or self.config.max_lag or 1)
        max_lag = int(max(1, max_lag))

        def _compute_at_lag(d: pd.DataFrame, lag: int) -> np.ndarray | None:
            return compute_connectivity_variant(
                d,
                variant,
                lag=int(max(1, lag)),
                control=kwargs.get("control"),
                pairs=pairs_idx,
                partial_mode=kwargs.get("partial_mode", "pairwise"),
                pairwise_policy=kwargs.get("pairwise_policy", "others"),
                custom_controls=kwargs.get("custom_controls"),
                control_strategy=kwargs.get("control_strategy", "none"),
                control_pca_k=int(kwargs.get("control_pca_k", 0) or 0),
            )

        chosen_lag = 1
        if not supports_lag:
            chosen_lag = 1
        elif lag_sel == "fixed":
            chosen_lag = int(max(1, kwargs.get("lag", 1)))
        else:
            # optimize
            best_score = float("-inf")
            best_lag = 1
            for lag in range(1, max_lag + 1):
                mat = _compute_at_lag(df, lag)
                score = _lag_quality(variant, mat)
                if np.isfinite(score) and float(score) > best_score:
                    best_score = float(score)
                    best_lag = int(lag)
            chosen_lag = int(best_lag)
            meta["lag_optimization"] = {
                "max_lag": int(max_lag),
                "criterion": "mean(|offdiag|)" if not is_pvalue_method(variant) else "mean(-log10(p))",
            }
        meta["chosen_lag"] = int(chosen_lag)

        # --- Диагностические сканы по окнам/лагам/позициям ---
        scans = {
            "window_pos": bool(kwargs.get("scan_window_pos", False)),
            "window_size": bool(kwargs.get("scan_window_size", False)),
            "lag": bool(kwargs.get("scan_lag", False)),
            "cube": bool(kwargs.get("scan_cube", False)),
        }

        w_start_min = kwargs.get("window_start_min")
        w_start_max = kwargs.get("window_start_max")
        w_stride = kwargs.get("window_stride")
        w_max_windows = int(kwargs.get("window_max_windows", 250))

        lag_grid = kwargs.get("lag_grid")
        if lag_grid is None:
            lmin = max(1, int(kwargs.get("lag_min", 1)))
            lmax = max(lmin, int(kwargs.get("lag_max", max_lag)))
            lstep = max(1, int(kwargs.get("lag_step", 1)))
            lag_grid = list(range(lmin, lmax + 1, lstep))
        else:
            try:
                lag_grid = [int(x) for x in lag_grid if int(x) >= 1]
            except Exception:
                lag_grid = [1]
            if not lag_grid:
                lag_grid = [1]

        window_sizes_grid = kwargs.get("window_sizes_grid")
        if window_sizes_grid is not None:
            try:
                window_sizes_grid = [int(x) for x in window_sizes_grid if int(x) >= 2]
            except Exception:
                window_sizes_grid = []
        else:
            window_sizes_grid = []

        if any(scans.values()):
            scan_meta: dict = {"flags": scans}
            ws_fallback = kwargs.get("window_sizes") or self.config.window_sizes or []
            ws_fallback = [int(w) for w in ws_fallback if int(w) >= 2]
            ws_list = window_sizes_grid if window_sizes_grid else ws_fallback

            # Если список размеров окна пуст — можно собрать его из window_min/max/step.
            if not ws_list:
                try:
                    wmin = int(kwargs.get("window_min", 0) or 0)
                    wmax = int(kwargs.get("window_max", 0) or 0)
                    wstep = int(kwargs.get("window_step", 0) or 0)
                    if wmin >= 2 and wmax >= wmin and wstep >= 1:
                        ws_list = list(range(wmin, wmax + 1, wstep))
                except Exception:
                    ws_list = []

            default_w = int(kwargs.get("window_size", ws_list[0] if ws_list else min(200, max(10, len(df) // 5))))
            default_w = int(max(2, min(default_w, len(df))))

            if scans["window_pos"]:
                stride = int(w_stride) if w_stride is not None else int(max(1, default_w // 5))
                info = analyze_sliding_windows_with_metric(
                    df, variant, window_size=default_w, stride=stride, lag=int(chosen_lag),
                    start_min=w_start_min, start_max=w_start_max, max_windows=w_max_windows,
                    return_matrices=True,
                    pairs=pairs_idx,
                )
                ticks = []
                for i, t in enumerate((info or {}).get("ticks") or []):
                    tid = f"pos_w{default_w}_l{int(chosen_lag)}_i{i}_s{int(t.get('start', 0))}"
                    ticks.append({"id": tid, **t})
                ext = (info or {}).get("extremes") or {}
                ext_ids = {
                    "best": ticks[ext.get("best")]["id"] if ticks and ext.get("best") is not None else None,
                    "median": ticks[ext.get("median")]["id"] if ticks and ext.get("median") is not None else None,
                    "worst": ticks[ext.get("worst")]["id"] if ticks and ext.get("worst") is not None else None,
                }
                scan_meta["window_pos"] = {
                    "window_size": default_w, "stride": stride, "lag": int(chosen_lag),
                    "curve": info.get("curve") if info else None,
                    "best_window": info.get("best_window") if info else None,
                    "ticks": ticks,
                    "extremes": ext_ids,
                }

            if scans["window_size"] and ws_list:
                xs, ys = [], []
                ticks = []
                for w in ws_list:
                    stride = int(w_stride) if w_stride is not None else int(max(1, int(w) // 5))
                    info = analyze_sliding_windows_with_metric(
                        df, variant, window_size=int(w), stride=stride, lag=int(chosen_lag),
                        start_min=w_start_min, start_max=w_start_max, max_windows=w_max_windows,
                        return_matrices=False,
                        pairs=pairs_idx,
                    )
                    bw = (info or {}).get("best_window") or {}
                    q = bw.get("metric", float("nan"))
                    xs.append(int(w))
                    ys.append(float(q) if np.isfinite(q) else float("nan"))
                    tid = f"size_w{int(w)}_l{int(chosen_lag)}"
                    ticks.append({
                        "id": tid,
                        "window_size": int(w),
                        "start": int(bw.get("start", 0)) if bw else 0,
                        "end": int(bw.get("end", 0)) if bw else 0,
                        "metric": float(q) if np.isfinite(q) else float("nan"),
                        "matrix": bw.get("matrix"),
                    })
                ext = _select_best_median_worst(ticks, key="metric")
                ext_ids = {
                    "best": ticks[ext.get("best")]["id"] if ticks and ext.get("best") is not None else None,
                    "median": ticks[ext.get("median")]["id"] if ticks and ext.get("median") is not None else None,
                    "worst": ticks[ext.get("worst")]["id"] if ticks and ext.get("worst") is not None else None,
                }
                scan_meta["window_size"] = {"lag": int(chosen_lag), "curve": {"x": xs, "y": ys}, "ticks": ticks, "extremes": ext_ids}

            if scans["lag"] and supports_lag:
                xs, ys = [], []
                ticks = []
                for lag in lag_grid:
                    mat_l = _compute_at_lag(df, int(lag))
                    q = _lag_quality(variant, mat_l)
                    xs.append(int(lag))
                    ys.append(float(q) if np.isfinite(q) else float("nan"))
                    tid = f"lag_l{int(lag)}"
                    ticks.append({"id": tid, "lag": int(lag), "metric": float(q) if np.isfinite(q) else float("nan"), "matrix": mat_l})
                ext = _select_best_median_worst(ticks, key="metric")
                ext_ids = {
                    "best": ticks[ext.get("best")]["id"] if ticks and ext.get("best") is not None else None,
                    "median": ticks[ext.get("median")]["id"] if ticks and ext.get("median") is not None else None,
                    "worst": ticks[ext.get("worst")]["id"] if ticks and ext.get("worst") is not None else None,
                }
                scan_meta["lag"] = {"curve": {"x": xs, "y": ys}, "grid": lag_grid, "ticks": ticks, "extremes": ext_ids}

            if scans["cube"] and ws_list:
                combo_limit = max(1, int(kwargs.get("cube_combo_limit", 80)))
                cube_eval_limit = int(kwargs.get("cube_eval_limit", 0) or 0)
                cube_matrix_mode = str(kwargs.get("cube_matrix_mode", "selected") or "selected").lower()
                if cube_matrix_mode not in {"selected", "all"}:
                    cube_matrix_mode = "selected"
                cube_matrix_limit = int(kwargs.get("cube_matrix_limit", 0) or 0)
                if cube_matrix_limit <= 0:
                    cube_matrix_limit = cube_eval_limit if cube_eval_limit > 0 else 500

                points: list[dict] = []
                lags_for_cube = lag_grid if supports_lag else [1]
                combos = [(int(w), int(lg)) for w in ws_list for lg in lags_for_cube]
                if len(combos) > combo_limit:
                    idx = np.linspace(0, len(combos) - 1, combo_limit).round().astype(int)
                    combos = [combos[i] for i in idx]

                per_combo_max_windows = int(w_max_windows)
                if cube_eval_limit and len(combos) > 0:
                    per_combo_max_windows = max(1, min(int(w_max_windows), int(cube_eval_limit) // int(len(combos))))
                saved_mats = 0

                for w, lg in combos:
                    stride = int(w_stride) if w_stride is not None else int(max(1, int(w) // 5))
                    info = analyze_sliding_windows_with_metric(
                        df, variant, window_size=w, stride=stride, lag=lg,
                        start_min=w_start_min, start_max=w_start_max,
                        max_windows=per_combo_max_windows,
                        return_matrices=(cube_matrix_mode == "all"),
                        pairs=pairs_idx,
                    )
                    for t in (info or {}).get("ticks") or []:
                        try:
                            fq = float(t.get("metric", float("nan")))
                        except Exception:
                            fq = float("nan")
                        if not np.isfinite(fq):
                            continue
                        st0 = int(t.get("start", 0))
                        tid = f"cube_w{int(w)}_l{int(lg)}_s{st0}"
                        mat0 = t.get("matrix")
                        if cube_matrix_mode == "all":
                            if mat0 is not None and saved_mats < cube_matrix_limit:
                                saved_mats += 1
                            else:
                                mat0 = None
                        points.append({
                            "id": tid,
                            "window_size": int(w), "lag": int(lg),
                            "start": st0, "end": int(t.get("end", st0 + int(w))),
                            "metric": fq,
                            "matrix": mat0,
                        })

                scan_meta["cube"] = {
                    "points": points, "combos": combos,
                    "lag_grid": lags_for_cube,
                    "window_sizes": ws_list,
                    "eval_limit": int(cube_eval_limit) if cube_eval_limit else None,
                    "matrix_mode": cube_matrix_mode,
                    "matrix_limit": int(cube_matrix_limit),
                }

                ext = _select_best_median_worst(points, key="metric")
                ext_ids = {
                    "best": points[ext.get("best")]["id"] if points and ext.get("best") is not None else None,
                    "median": points[ext.get("median")]["id"] if points and ext.get("median") is not None else None,
                    "worst": points[ext.get("worst")]["id"] if points and ext.get("worst") is not None else None,
                }
                scan_meta["cube"]["extremes"] = ext_ids
                if ext.get("best") is not None:
                    points[int(ext["best"])]["tag"] = "best"
                if ext.get("median") is not None:
                    points[int(ext["median"])]["tag"] = "median"
                if ext.get("worst") is not None:
                    points[int(ext["worst"])]["tag"] = "worst"

                if cube_matrix_mode == "all":
                    must = [ext.get("best"), ext.get("median"), ext.get("worst")]
                    for ii in must:
                        if ii is None:
                            continue
                        ii = int(ii)
                        if ii < 0 or ii >= len(points):
                            continue
                        if points[ii].get("matrix") is not None:
                            continue
                        try:
                            w0 = int(points[ii].get("window_size"))
                            lg0 = int(points[ii].get("lag"))
                            st0 = int(points[ii].get("start"))
                            chunk = df.iloc[st0 : st0 + w0]
                            points[ii]["matrix"] = compute_connectivity_variant(chunk, variant, lag=int(max(1, lg0)))
                        except Exception:
                            points[ii]["matrix"] = None

                # Матрицы для выбранных точек 3D-куба (best/median/worst + выборка).
                gallery_k = max(1, int(kwargs.get("cube_gallery_k", 1)))
                gallery_limit = max(3, int(kwargs.get("cube_gallery_limit", 60)))
                gallery_mode = str(kwargs.get("cube_gallery_mode", "extremes") or "extremes").lower()

                pts_sorted = [p for p in points if np.isfinite(float(p.get("metric", float("nan"))))]
                pts_sorted.sort(key=lambda p: float(p.get("metric", float("nan"))))
                if pts_sorted:
                    idx_set = {0, len(pts_sorted) // 2, len(pts_sorted) - 1}

                    if gallery_mode in {"topbottom", "extremes"}:
                        for i in range(gallery_k):
                            idx_set.add(min(i, len(pts_sorted) - 1))
                            idx_set.add(max(0, len(pts_sorted) - 1 - i))

                    if gallery_mode == "quantiles":
                        for q in (0.1, 0.25, 0.5, 0.75, 0.9):
                            idx_set.add(int(round(q * (len(pts_sorted) - 1))))

                    idx_list = sorted(idx_set)
                    if len(idx_list) > gallery_limit:
                        sel = np.linspace(0, len(idx_list) - 1, gallery_limit).round().astype(int)
                        idx_list = [idx_list[i] for i in sel]

                    gallery = []
                    for ii in idx_list:
                        p = pts_sorted[ii]
                        w0 = int(p.get("window_size"))
                        lg0 = int(p.get("lag"))
                        st0 = int(p.get("start"))
                        tid0 = p.get("id") or f"cube_w{w0}_l{lg0}_s{st0}"
                        try:
                            chunk = df.iloc[st0 : st0 + w0]
                            mat0 = compute_connectivity_variant(chunk, variant, lag=int(max(1, lg0)))
                        except Exception:
                            mat0 = None
                        gallery.append(
                            {
                                "id": tid0,
                                "window_size": w0,
                                "lag": lg0,
                                "start": st0,
                                "end": int(st0 + w0),
                                "metric": float(p.get("metric")),
                                "tag": p.get("tag"),
                                "matrix": mat0,
                            }
                        )

                    scan_meta["cube"]["gallery"] = gallery
                    scan_meta["cube"]["selectable_ids"] = [g.get("id") for g in gallery if g.get("matrix") is not None]
                else:
                    scan_meta["cube"]["selectable_ids"] = []

                if cube_matrix_mode == "all":
                    scan_meta["cube"]["selectable_ids"] = [p.get("id") for p in points if p.get("matrix") is not None]

                # --- дополнительные "кубики" по парам (только если ровно 3 переменные
                # или пользователь явно передал список пар через kwargs).
                # Идея: один и тот же набор точек (w,lag,start) визуализируем разными метриками
                # для пар (X–Y, X–Z, Y–Z).
                try:
                    cube_pairs = kwargs.get("cube_pairs")
                except Exception:
                    cube_pairs = None

                pair_specs: list[tuple[int, int, str]] = []
                cols0 = list(getattr(df, "columns", []))
                if cube_pairs:
                    # ожидаем список вида [(0,1), ("X","Y"), "X-Y", ...]
                    for item in (cube_pairs or []):
                        try:
                            if isinstance(item, (list, tuple)) and len(item) >= 2:
                                a, b = item[0], item[1]
                            elif isinstance(item, str):
                                s = item.replace("→", "-").replace(">", "-").replace(":", "-")
                                a, b = [x.strip() for x in s.split("-")[:2]]
                            else:
                                continue

                            if isinstance(a, int) and isinstance(b, int):
                                i, j = int(a), int(b)
                                name = f"{cols0[i]}—{cols0[j]}" if i < len(cols0) and j < len(cols0) else f"{i}—{j}"
                                pair_specs.append((i, j, name))
                            else:
                                if str(a) in cols0 and str(b) in cols0:
                                    i, j = cols0.index(str(a)), cols0.index(str(b))
                                    pair_specs.append((i, j, f"{a}—{b}"))
                        except Exception:
                            continue
                elif int(df.shape[1]) == 3:
                    pair_specs = [
                        (0, 1, f"{cols0[0]}—{cols0[1]}"),
                        (0, 2, f"{cols0[0]}—{cols0[2]}"),
                        (1, 2, f"{cols0[1]}—{cols0[2]}"),
                    ]

                if pair_specs:
                    cubes_by_pair: dict[str, dict] = {}
                    for i, j, nm in pair_specs:
                        pts_p = []
                        for p in points:
                            pid = p.get("id")
                            mat0 = p.get("matrix")
                            if pid is None or mat0 is None:
                                continue
                            scv = _pair_score(variant, np.asarray(mat0), int(i), int(j))
                            if not np.isfinite(float(scv)):
                                continue
                            pts_p.append({
                                "id": pid,
                                "window_size": p.get("window_size"),
                                "lag": p.get("lag"),
                                "start": p.get("start"),
                                "end": p.get("end"),
                                "metric": float(scv),
                                "tag": p.get("tag"),
                            })

                        extp = _select_best_median_worst(pts_p, key="metric")
                        cubes_by_pair[nm] = {
                            "pair": [int(i), int(j)],
                            "points": pts_p,
                            "extremes": {
                                "best": (pts_p[int(extp["best"])]["id"] if pts_p and extp.get("best") is not None else None),
                                "median": (pts_p[int(extp["median"])]["id"] if pts_p and extp.get("median") is not None else None),
                                "worst": (pts_p[int(extp["worst"])]["id"] if pts_p and extp.get("worst") is not None else None),
                            },
                        }

                    if cubes_by_pair:
                        scan_meta["cube_pairs"] = cubes_by_pair

            meta["window_scans"] = scan_meta

        # --- windowing ---
        window_sizes = kwargs.get("window_sizes") or self.config.window_sizes
        if not window_sizes:
            mat = _compute_at_lag(df, chosen_lag)
            return mat, meta

        policy = (kwargs.get("window_policy") or self.config.window_policy or "best").lower()
        window_sizes = [int(w) for w in window_sizes if int(w) >= 2]
        stride_override = kwargs.get("window_stride") or self.config.window_stride

        window_cube_level = str(kwargs.get("window_cube_level", "off") or "off").lower()
        if window_cube_level != "off":
            # Быстрый 3D-поиск: window_size × lag × window_start.
            max_lag_cube = int(kwargs.get("max_lag", self.config.max_lag) or self.config.max_lag)
            lags = list(range(1, max(1, max_lag_cube) + 1))
            eval_limit = int(kwargs.get("window_cube_eval_limit", 120 if window_cube_level == "full" else 60))
            grid = []
            best_cube = None
            best_cube_q = float("-inf")

            combos = [(w, lag) for w in window_sizes for lag in lags]
            if len(combos) > eval_limit:
                idx = np.linspace(0, len(combos) - 1, eval_limit).round().astype(int)
                combos = [combos[i] for i in idx]

            for w, lag in combos:
                stride = int(stride_override) if stride_override is not None else int(max(1, int(w) // 5))
                w_info = analyze_sliding_windows_with_metric(df, variant, window_size=int(w), stride=int(stride), lag=int(lag), pairs=pairs_idx)
                bw = w_info.get("best_window") if w_info else None
                q = float(bw.get("metric", float("nan"))) if bw else float("nan")
                grid.append({"window_size": int(w), "lag": int(lag), "best_metric": q, "best_start": (bw.get("start") if bw else None)})
                if np.isfinite(q) and q > best_cube_q and bw and bw.get("matrix") is not None:
                    best_cube_q = q
                    best_cube = {
                        "window_size": int(w),
                        "lag": int(lag),
                        "stride": int(stride),
                        "best_window": bw,
                        "curve": (w_info.get("curve") if window_cube_level == "full" else None),
                    }

            if best_cube is not None:
                # Для отчёта сохраняем и 3D-точки (window, lag, start, metric), и агрегированный grid.
                points = []
                try:
                    for g in grid:
                        st = g.get("best_start")
                        mt = g.get("best_metric")
                        if st is None or mt is None:
                            continue
                        points.append({"window_size": int(g["window_size"]), "lag": int(g["lag"]), "start": int(st), "metric": float(mt)})
                except Exception:
                    points = []

                meta["window_cube"] = {
                    "level": window_cube_level,
                    "eval_limit": int(eval_limit),
                    "grid": grid,
                    "points": points,
                    "best": best_cube,
                }
                chosen_lag = int(best_cube["lag"])

        mats = []
        best = None
        best_q = float("-inf")

        for w in window_sizes:
            stride = int(stride_override) if stride_override is not None else int(max(1, w // 5))
            w_info = analyze_sliding_windows_with_metric(df, variant, window_size=w, stride=stride, lag=chosen_lag, pairs=pairs_idx)
            if not w_info:
                continue
            # best window for this w
            bw = w_info.get("best_window")
            if bw and bw.get("matrix") is not None:
                mats.append(np.asarray(bw["matrix"]))
                q = float(bw.get("metric", float("nan")))
                if np.isfinite(q) and q > best_q:
                    best_q = q
                    best = {
                        "window_size": int(w),
                        "stride": int(stride),
                        "best_window": bw,
                        "curve": w_info.get("curve"),
                    }

        if not mats:
            mat = _compute_at_lag(df, chosen_lag)
            return mat, meta

        if policy == "mean":
            # усредняем по матрицам лучших окон каждого размера
            mat = np.nanmean(np.stack(mats, axis=0), axis=0)
        else:
            # best
            mat = np.asarray(best["best_window"]["matrix"]) if best else np.asarray(mats[0])

        meta["window"] = {
            "sizes": window_sizes,
            "policy": policy,
            "best": best,
        }
        return mat, meta

    def export_html_report(self, output_path: str, **kwargs) -> str:
        """Генерация HTML отчета через внешний класс."""
        return HTMLReportGenerator(self).generate(output_path, **kwargs)

    def export_big_excel(self, save_path: str, **kwargs) -> str:
        """Генерация Excel отчета через внешний класс."""
        return ExcelReportWriter(self).write(save_path, **kwargs)

    def export_series_bundle(self, save_path: str) -> str:
        """Сохраняет сами ряды (RAW/после предобработки/после auto-diff/normalized) единым файлом.

        Это отдельный файл, который удобно держать рядом с отчётами (HTML/Excel).
        """
        import pandas as pd

        def _pick(df):
            try:
                return df if df is not None and not getattr(df, 'empty', True) else None
            except Exception:
                return None

        raw_df = _pick(getattr(self, 'data_raw', None)) or _pick(getattr(self, 'data', None)) or pd.DataFrame()
        pre_df = _pick(getattr(self, 'data_preprocessed', None)) or _pick(getattr(self, 'data', None)) or pd.DataFrame()
        ad_df = _pick(getattr(self, 'data_after_autodiff', None)) or _pick(getattr(self, 'data', None)) or pd.DataFrame()
        norm_df = _pick(getattr(self, 'data_normalized', None)) or pd.DataFrame()

        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            raw_df.to_excel(writer, sheet_name='RAW', index=False)
            pre_df.to_excel(writer, sheet_name='PREPROCESSED', index=False)
            ad_df.to_excel(writer, sheet_name='AFTER_AUTODIFF', index=False)
            if not norm_df.empty:
                norm_df.to_excel(writer, sheet_name='NORMALIZED', index=False)

            # QC (если есть)
            try:
                qc_raw = getattr(self, 'qc_raw', None)
                qc_clean = getattr(self, 'qc_clean', None)
                if qc_raw is not None and not getattr(qc_raw, 'empty', True):
                    qc_raw.to_excel(writer, sheet_name='QC_RAW', index=False)
                if qc_clean is not None and not getattr(qc_clean, 'empty', True):
                    qc_clean.to_excel(writer, sheet_name='QC_CLEAN', index=False)
            except Exception:
                pass

            # Координаты (если есть)
            try:
                coords = getattr(self, 'coords_df', None)
                if coords is not None and not getattr(coords, 'empty', True):
                    coords.to_excel(writer, sheet_name='COORDS', index=False)
            except Exception:
                pass

        logging.info('[Series] Сохранены ряды: %s', save_path)

        # Метаданные прогона рядом (JSON)
        try:
            import json
            meta_path = Path(save_path).with_suffix('.meta.json')
            meta = {
                'series_file': str(Path(save_path).name),
                'data_shape': list(getattr(self, 'data', pd.DataFrame()).shape),
                'preprocessing': self.get_preprocessing_summary(),
                'methods': list(getattr(self, 'results', {}).keys()),
                'results_meta': getattr(self, 'results_meta', {}),
            }
            meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')
        except Exception:
            pass
        return save_path

    def _maybe_apply_dimred(self, **kwargs) -> None:
        """Предобработка: опциональное уменьшение размерности для больших N."""
        enabled = bool(kwargs.get("dimred_enabled", False))
        method = str(kwargs.get("dimred_method") or "none").strip().lower()
        target_n = int(kwargs.get("dimred_target", 0) or 0)
        target_var = kwargs.get("dimred_target_var", None)
        try:
            target_var = float(target_var) if target_var is not None and str(target_var).strip() != "" else None
        except Exception:
            target_var = None
        save_variants = bool(kwargs.get("dimred_save_variants", False))
        variants_text = str(kwargs.get("dimred_variants") or "").strip()
        seed = int(kwargs.get("dimred_seed", 0) or 0)
        spatial_bin = int(kwargs.get("dimred_spatial_bin", 2) or 2)
        kmeans_batch = int(kwargs.get("dimred_kmeans_batch", 2048) or 2048)

        base = (
            self.data_preprocessed
            if isinstance(getattr(self, "data_preprocessed", None), pd.DataFrame) and not self.data_preprocessed.empty
            else getattr(self, "data", None)
        )
        if base is None or getattr(base, "empty", True):
            self.dimred_report = {"enabled": False, "reason": "no_data"}
            self.data_dimred = pd.DataFrame()
            self.dimred_mapping = pd.DataFrame()
            return

        self.data_dimred_base = base
        n0 = int(base.shape[1])
        if (not enabled) or (method in ("none", "off", "disabled")):
            self.dimred_report = {"enabled": False, "method": "none", "k": n0}
            self.data_dimred = base.copy()
            self.dimred_mapping = pd.DataFrame({"source": list(base.columns), "target": list(base.columns), "weight": 1.0})
            self.data = base.copy()
            return

        if (target_n <= 0) and (target_var is None or not (0.0 < target_var <= 1.0)):
            target_n = int(min(500, n0))

        res = apply_dimred(
            base,
            method=method,
            target_n=(int(min(target_n, n0)) if target_n and target_n > 0 else None),
            target_var=target_var,
            seed=seed,
            coords_df=getattr(self, "coords_df", None),
            kmeans_batch=kmeans_batch,
            spatial_bin=spatial_bin,
        )
        self.data_dimred = res.reduced
        self.dimred_mapping = res.mapping
        self.dimred_report = {"enabled": True, **(res.meta or {}), "n_before": n0, "n_after": int(res.reduced.shape[1])}
        self.data_preprocessed = self.data_dimred.copy()
        self.data = self.data_dimred.copy()

        if save_variants and variants_text:
            try:
                parts = [p.strip() for p in variants_text.replace(";", ",").split(",") if p.strip()]
                targets = sorted(set(int(float(p)) for p in parts if float(p) > 0))
                self.dimred_report["saved_variants"] = targets
            except Exception as e:
                self.dimred_report["saved_variants_error"] = str(e)

    def export_dimred_bundle(self, out_dir: str, name_prefix: str = "run") -> Dict[str, str]:
        """Экспортирует уменьшенные ряды и source->target mapping в ``out_dir/data``."""
        import json
        from pathlib import Path

        data_dir = Path(out_dir) / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        paths: Dict[str, str] = {}
        pref = f"{name_prefix}_" if name_prefix else ""

        try:
            df = getattr(self, "data_dimred", None)
            if df is not None and not getattr(df, "empty", True):
                p = data_dir / f"{pref}timeseries_dimred.csv"
                df.to_csv(p, index=True)
                paths["timeseries_dimred_csv"] = str(p)
        except Exception:
            pass

        try:
            mp = getattr(self, "dimred_mapping", None)
            if mp is not None and not getattr(mp, "empty", True):
                p = data_dir / f"{pref}dimred_mapping.csv"
                mp.to_csv(p, index=False)
                paths["dimred_mapping_csv"] = str(p)
        except Exception:
            pass

        try:
            rep = getattr(self, "dimred_report", {}) or {}
            p = data_dir / f"{pref}dimred_meta.json"
            p.write_text(json.dumps(rep, ensure_ascii=False, indent=2), encoding="utf-8")
            paths["dimred_meta_json"] = str(p)
        except Exception:
            pass

        try:
            rep = getattr(self, "dimred_report", {}) or {}
            targets = rep.get("saved_variants") or []
            m = str(rep.get("method", "none")).strip().lower()
            if targets and m not in ("none", "off", "disabled"):
                base = getattr(self, "data_dimred_base", None)
                if base is not None and not getattr(base, "empty", True):
                    vroot = data_dir / f"{pref}dimred_variants"
                    vroot.mkdir(parents=True, exist_ok=True)
                    for t in targets:
                        try:
                            t2 = int(min(int(t), int(base.shape[1])))
                            rr = apply_dimred(
                                base,
                                method=m,
                                target_n=t2,
                                seed=int(rep.get("seed", 0) or 0),
                                coords_df=getattr(self, "coords_df", None),
                                kmeans_batch=int(rep.get("batch_size", 2048) or 2048),
                                spatial_bin=int(rep.get("bin_size", 2) or 2),
                            )
                            sub = vroot / f"{m}_n{t2}"
                            sub.mkdir(parents=True, exist_ok=True)
                            rr.reduced.to_csv(sub / "timeseries_dimred.csv", index=True)
                            rr.mapping.to_csv(sub / "dimred_mapping.csv", index=False)
                            (sub / "dimred_meta.json").write_text(json.dumps(rr.meta, ensure_ascii=False, indent=2), encoding="utf-8")
                        except Exception:
                            continue
                    paths["dimred_variants_dir"] = str(vroot)
        except Exception:
            pass

        return paths

    def _maybe_post_preprocess(self, **kwargs) -> None:
        """Опциональная предобработка после dimred (для stage=post|both)."""
        stage = str(kwargs.get("preprocess_stage", "pre")).strip().lower()
        if stage not in ("post", "both"):
            return

        post = kwargs.get("post_preprocess", {}) or {}
        if not bool(post.get("enabled", False)):
            return

        try:
            self.data = preprocess_timeseries(
                self.data,
                enabled=True,
                log_transform=bool(post.get("log_transform", False)),
                remove_outliers=bool(post.get("remove_outliers", False)),
                outlier_rule=str(post.get("outlier_rule", "robust_z")),
                outlier_action=str(post.get("outlier_action", "mask")),
                outlier_z=float(post.get("outlier_z", 5.0)),
                outlier_k=float(post.get("outlier_k", 1.5)),
                outlier_abs=post.get("outlier_abs", None),
                outlier_p_low=float(post.get("outlier_p_low", 0.5)),
                outlier_p_high=float(post.get("outlier_p_high", 99.5)),
                outlier_hampel_window=int(post.get("outlier_hampel_window", 7)),
                outlier_jump_thr=post.get("outlier_jump_thr", None),
                outlier_local_median_window=int(post.get("outlier_local_median_window", 7)),
                normalize=bool(post.get("normalize", True)),
                normalize_mode=str(post.get("normalize_mode", "zscore")),
                rank_mode=str(post.get("rank_mode", "dense")),
                rank_ties=str(post.get("rank_ties", "average")),
                fill_missing=bool(post.get("fill_missing", True)),
                remove_ar1=bool(post.get("remove_ar1", False)),
                remove_ar_order=int(post.get("remove_ar_order", 1) or 1),
                remove_seasonality=bool(post.get("remove_seasonality", False)),
                season_period=post.get("season_period", None),
                check_stationarity=False,
                return_report=False,
            )
            try:
                getattr(self, "log", RunLog()).add("Post-preprocess: applied")
            except Exception:
                logging.info("Post-preprocess: applied")
        except Exception as e:
            try:
                getattr(self, "log", RunLog()).add(f"Post-preprocess failed: {e}")
            except Exception:
                logging.warning("Post-preprocess failed: %s", e)


    def export_connectivity_bundle(
        self,
        out_dir: str,
        name_prefix: str = "run",
        dense_n_limit: int = 2000,
        topk_per_node: int = 50,
        min_abs_weight: float = 0.0,
        include_scan_matrices: bool = True,
    ) -> str:
        """Экспортирует матрицы связности и графы как данные для внешних пайплайнов.

        Записывает в ``out_dir/data``:
        - ``nodes.csv`` с узлами и (опционально) координатами,
        - для каждой матрицы: ``edges.csv.gz``, ``sparse.npz`` и ``dense.npy`` (только для малых N),
        - ``manifest.json`` со сводной метаинформацией.

        Для больших размерностей плотная NxN-матрица может быть тяжёлой,
        поэтому sparse-представление сохраняется всегда.
        """
        from pathlib import Path

        import numpy as np

        from src.reporting.connectivity_export import ExportPolicy, export_connectivity_matrix, save_manifest, save_nodes_csv

        data_dir = str(Path(out_dir) / "data")
        os.makedirs(data_dir, exist_ok=True)

        # Экспорт артефактов уменьшения размерности (если использовалось).
        try:
            self.export_dimred_bundle(out_dir, name_prefix=name_prefix)
        except Exception:
            pass

        # Имена узлов из доступных источников.
        try:
            names = list(getattr(self, "node_names", None) or getattr(self, "columns", None) or [])
        except Exception:
            names = []
        if not names:
            try:
                names = list(getattr(self, "data", None).columns)
            except Exception:
                names = []

        # Координаты (если загружались вместе с данными).
        coords_map = None
        try:
            coords_df = getattr(self, "coords_df", None)
            if coords_df is not None and not getattr(coords_df, "empty", True):
                ccols = [c.lower() for c in coords_df.columns]
                name_col = "name" if "name" in ccols else ("node" if "node" in ccols else None)
                if name_col:
                    coords_map = {}
                    for _, row in coords_df.iterrows():
                        nm = str(row[name_col])
                        x = row.get("x", row.get("X", None))
                        y = row.get("y", row.get("Y", None))
                        z = row.get("z", row.get("Z", None))
                        coords_map[nm] = (x, y, z)
        except Exception:
            coords_map = None

        policy = ExportPolicy(
            dense_n_limit=int(dense_n_limit),
            topk_per_node=int(topk_per_node),
            min_abs_weight=float(min_abs_weight),
        )

        manifest: dict = {
            "name_prefix": str(name_prefix),
            "policy": {
                "dense_n_limit": int(dense_n_limit),
                "topk_per_node": int(topk_per_node),
                "min_abs_weight": float(min_abs_weight),
            },
            "variants": {},
            "scan_matrices": {},
        }

        if names:
            save_nodes_csv(data_dir, names, coords=coords_map)
            manifest["nodes_file"] = "nodes.csv"

        # Основные матрицы по выбранным методам.
        results = getattr(self, "results", {}) or {}
        for variant, mat in results.items():
            try:
                arr = np.asarray(mat)
                if arr.ndim != 2 or arr.shape[0] != arr.shape[1]:
                    continue
                if names and arr.shape[0] != len(names):
                    # Размер не совпал: используем локальные имена по индексу.
                    local_names = [f"v{i:04d}" for i in range(arr.shape[0])]
                else:
                    local_names = names or [f"v{i:04d}" for i in range(arr.shape[0])]
                extra = None
                try:
                    extra = (getattr(self, "results_meta", {}) or {}).get(variant)
                except Exception:
                    extra = None
                m = export_connectivity_matrix(data_dir, str(name_prefix), str(variant), arr, local_names, policy, extra_meta=extra)
                manifest["variants"][variant] = m
            except Exception as e:
                manifest["variants"][variant] = {"error": str(e)}

        # Матрицы из сканов (если доступны в results_meta).
        if include_scan_matrices:
            try:
                scans = (getattr(self, "results_meta", {}) or {}).get("scans", {})
            except Exception:
                scans = {}
            if isinstance(scans, dict):
                for scan_name, payload in scans.items():
                    if not isinstance(payload, dict):
                        continue
                    for key in ("best", "median", "worst"):
                        item = payload.get(key)
                        if not isinstance(item, dict):
                            continue
                        mat = item.get("matrix")
                        if mat is None:
                            continue
                        try:
                            arr = np.asarray(mat)
                            if arr.ndim != 2 or arr.shape[0] != arr.shape[1]:
                                continue
                            local_names = names or [f"v{i:04d}" for i in range(arr.shape[0])]
                            vname = f"scan_{scan_name}_{key}"
                            m = export_connectivity_matrix(data_dir, str(name_prefix), vname, arr, local_names, policy, extra_meta=item)
                            manifest["scan_matrices"].setdefault(scan_name, {})[key] = m
                        except Exception as e:
                            manifest["scan_matrices"].setdefault(scan_name, {})[key] = {"error": str(e)}

        manifest_path = save_manifest(data_dir, str(name_prefix), manifest)
        return str(manifest_path)

    def test_stationarity(self, series: pd.Series) -> Tuple[Optional[float], Optional[float]]:
        """Проверяет стационарность ряда через ADF-тест."""
        return analysis_stats.test_stationarity(series)

    def get_preprocessing_summary(self) -> dict:
        """Возвращает отчёт о предобработке и auto-diff в формате для UI/HTML."""
        rep = {}
        try:
            if self.preprocessing_report is not None:
                pr = self.preprocessing_report
                rep["preprocess"] = {
                    "enabled": bool(getattr(pr, "enabled", True)),
                    "steps_global": list(getattr(pr, "steps_global", [])),
                    "steps_by_column": dict(getattr(pr, "steps_by_column", {})),
                    "dropped_columns": list(getattr(pr, "dropped_columns", [])),
                    "notes": dict(getattr(pr, "notes", {})),
                }
            else:
                rep["preprocess"] = {"enabled": None, "steps_global": [], "steps_by_column": {}, "dropped_columns": [], "notes": {}}
        except Exception:
            rep["preprocess"] = {"enabled": None, "steps_global": [], "steps_by_column": {}, "dropped_columns": [], "notes": {}}

        rep["autodiff"] = dict(self.autodiff_report or {"enabled": False, "differenced": []})
        return rep

    def get_harmonics(self, top_k: int = 5, fs: float | None = None) -> dict:
        """Возвращает FFT-пики (гармоники) по каждому ряду."""
        out = {}
        if self.data.empty:
            return out
        fs0 = float(fs) if fs is not None else float(getattr(self, "fs", 1.0))
        for col in self.data.columns:
            s = self.data[col]
            out[col] = analysis_stats.fft_peaks(s, fs=fs0, top_k=int(max(1, top_k)))
        return out

    def get_diagnostics(self) -> dict:
        """Возвращает базовые диагностические метрики по всем столбцам."""
        diagnostics = {}
        if self.data.empty:
            return diagnostics
        for col in self.data.columns:
            series = self.data[col]
            adf_stat, adf_p = analysis_stats.test_stationarity(series)
            season = analysis_stats.detect_seasonality(series)
            fft_pk = analysis_stats.fft_peaks(series, top_k=3)
            diagnostics[col] = {
                "adf_stat": adf_stat,
                "adf_p": adf_p,
                "hurst_rs": analysis_stats.compute_hurst_rs(series),
                "hurst_dfa": analysis_stats.compute_hurst_dfa(series),
                "hurst_aggvar": analysis_stats.compute_hurst_aggvar(series),
                "hurst_wavelet": analysis_stats.compute_hurst_wavelet(series),
                "sample_entropy": analysis_stats.compute_sample_entropy(series),
                "shannon_entropy": analysis_stats.shannon_entropy(series),
                "permutation_entropy": analysis_stats.permutation_entropy(series, order=3, delay=1, normalize=True),
                "seasonality": season,
                "fft_peaks": fft_pk,
            }
        return diagnostics

    def build_pairwise_summaries(self, *, p_alpha: float = 0.05) -> None:
        """Строит компактные pairwise-таблицы по каждой матрице результатов."""
        import pandas as pd

        self.pairwise_summaries = {}
        cols = list(self.data.columns) if self.data is not None and not self.data.empty else []
        for variant, mat in (self.results or {}).items():
            if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
                continue
            is_pval = is_pvalue_method(variant)
            thr = float(p_alpha) if is_pval else float(getattr(self.config, "graph_threshold", 0.2))
            rows = []
            for i, a in enumerate(cols):
                for j, b in enumerate(cols):
                    if i == j:
                        continue
                    v = mat[i, j]
                    if v is None or not np.isfinite(float(v)):
                        continue
                    if is_pval:
                        flag = "significant" if float(v) < thr else ""
                    else:
                        flag = "strong" if abs(float(v)) > thr else ""
                    rows.append({"src": a, "tgt": b, "value": float(v), "flag": flag})
            if rows:
                self.pairwise_summaries[variant] = pd.DataFrame(rows)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Compute connectivity measures for multivariate time series."
    )
    parser.add_argument(
        "input_file",
        help="Path to input CSV or Excel file with time series data",
    )
    parser.add_argument(
        "--lags",
        type=int,
        default=DEFAULT_MAX_LAG,
        help="Lag or model order (for Granger, TE, etc.)",
    )
    parser.add_argument(
        "--pvalue-alpha",
        type=float,
        default=DEFAULT_PVALUE_ALPHA,
        help="Alpha for p-value methods (Granger full/directed)",
    )
    parser.add_argument(
        "--log",
        action="store_true",
        help="Apply logarithm transform to data (for positive-valued data)",
    )
    parser.add_argument(
        "--no-outliers",
        action="store_true",
        help="Disable outlier removal",
    )
    parser.add_argument(
        "--no-normalize",
        action="store_true",
        help="Disable normalization of data",
    )
    parser.add_argument(
        "--no-stationarity-check",
        action="store_true",
        help="Disable stationarity check (ADF test)",
    )
    parser.add_argument(
        "--graph-threshold",
        type=float,
        default=0.5,
        help="Threshold for graph edges (weight >= threshold)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel file path (defaults to TimeSeriesAnalysis/AllMethods_Full.xlsx)",
    )
    parser.add_argument(
        "--quiet-warnings",
        action="store_true",
        help="Suppress most warnings for cleaner CLI output.",
    )
    parser.add_argument(
        "--experimental",
        action="store_true",
        help="Enable experimental sliding-window analyses.",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Skip Excel report generation.",
    )
    parser.add_argument(
        "--report-html",
        default=None,
        dest="report_html",
        help="Path for HTML report output (optional).",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    configure_warnings(quiet=args.quiet_warnings)

    filepath = os.path.abspath(args.input_file)
    output_path = args.output or os.path.join(SAVE_FOLDER, "AllMethods_Full.xlsx")
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    tool = BigMasterTool(enable_experimental=args.experimental)
    tool.lag_ranges = {v: range(1, args.lags + 1) for v in method_mapping}
    tool.load_data_excel(
        filepath,
        log_transform=args.log,
        remove_outliers=not args.no_outliers,
        normalize=not args.no_normalize,
        fill_missing=True,
        check_stationarity=not args.no_stationarity_check,
    )
    tool.run_all_methods()
    do_excel = not args.no_excel
    do_report = bool(args.report_html)
    if not do_excel and not do_report:
        do_excel = True

    if do_excel:
        tool.export_big_excel(
            output_path,
            threshold=args.graph_threshold,
            p_value_alpha=args.pvalue_alpha,
        )

    if do_report:
        report_path = os.path.abspath(args.report_html)
        report_dir = os.path.dirname(report_path)
        if report_dir:
            os.makedirs(report_dir, exist_ok=True)
        tool.export_html_report(
            report_path,
            graph_threshold=args.graph_threshold,
            p_value_alpha=args.pvalue_alpha,
        )

    print("Анализ завершён, результаты сохранены в:", output_path)
